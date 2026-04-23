# -*- coding: utf-8 -*-
"""
KPI Type Classification and Type-Based Anomaly Detection Module

KPI Types (based on anomaly detection strategy):
- ceiling_saturated: Success rate ~100%, σ=0 → ABS_THRESHOLD ONLY
- floor_saturated: Failure count ~0, σ=0 → ABS_THRESHOLD ONLY  
- count_type: Discrete counts, high variance → RANGE + CONTEXT
- ratio_type: Non-saturated 0-100% → RELATIVE CHANGE
- physical_stable: dB/dBm, low variance → ABS_DELTA + TREND
- throughput_type: MB/bytes, right-skewed → QUANTILES
- latency_type: ms/us, right-skewed → TAIL AWARE
- low_frequency: Mostly 0, occasional 1 → WINDOW COUNT
- config_type: Discrete states → SEMANTIC RULE
- unknown: Default → STANDARD SPC
"""

import re
import json
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from enum import Enum
import xml.etree.ElementTree as ET


class KPIType(Enum):
    """KPI classification types based on anomaly detection strategy"""
    CEILING_SATURATED = "ceiling_saturated"
    FLOOR_SATURATED = "floor_saturated"
    COUNT_TYPE = "count_type"
    RATIO_TYPE = "ratio_type"
    PHYSICAL_STABLE = "physical_stable"
    THROUGHPUT_TYPE = "throughput_type"
    LATENCY_TYPE = "latency_type"
    LOW_FREQUENCY = "low_frequency"
    CONFIG_TYPE = "config_type"
    UNKNOWN = "unknown"


class DetectionStrategy(Enum):
    """Anomaly detection strategies (ivy0205 cleaned)
    
    7 Core strategies:
    - STANDARD_SPC: Default σ-based SPC rules (normal distribution)
    - ABS_THRESHOLD: Absolute threshold for ceiling/floor saturated data
    - IQR: Interquartile range, robust to outliers and non-normal data
    - POISSON: Poisson distribution for count/discrete data
    - RELATIVE: Relative percentage change for ratio data  
    - Z_SCORE: Z-score based detection for moderate variance
    - KS_TEST: Kolmogorov-Smirnov test for distribution comparison
    
    Legacy strategies are deprecated but kept for backward compatibility.
    Use normalize_strategy() to convert them to core strategies.
    """
    # Core strategies (7 valid)
    STANDARD_SPC = "STANDARD_SPC"   # Default: σ-based SPC for normal data
    ABS_THRESHOLD = "ABS_THRESHOLD" # Absolute threshold for saturated data
    IQR = "IQR"                     # IQR-based for high-variance/non-normal
    POISSON = "POISSON"             # Poisson distribution for count data
    RELATIVE = "RELATIVE"           # Relative % change for ratio data
    Z_SCORE = "Z_SCORE"             # Z-score for moderate variance
    KS_TEST = "KS_TEST"             # K-S test for distribution comparison
    
    # Legacy strategies (deprecated, kept for backward compatibility only)
    # These are automatically converted to core strategies via normalize_strategy()
    SKIP = "SKIP"                   # → ABS_THRESHOLD
    RANGE_CONTEXT = "RANGE_CONTEXT"       # → IQR
    RELATIVE_CHANGE = "RELATIVE_CHANGE"   # → RELATIVE
    ABS_DELTA_TREND = "ABS_DELTA_TREND"   # → Z_SCORE
    QUANTILES = "QUANTILES"               # → IQR
    TAIL_AWARE = "TAIL_AWARE"             # → IQR
    WINDOW_COUNT = "WINDOW_COUNT"         # → POISSON
    SEMANTIC_RULE = "SEMANTIC_RULE"       # → POISSON


# Strategy mapping: Legacy → Core
LEGACY_STRATEGY_MAP = {
    DetectionStrategy.SKIP: DetectionStrategy.ABS_THRESHOLD,
    DetectionStrategy.RANGE_CONTEXT: DetectionStrategy.IQR,
    DetectionStrategy.RELATIVE_CHANGE: DetectionStrategy.RELATIVE,
    DetectionStrategy.ABS_DELTA_TREND: DetectionStrategy.Z_SCORE,
    DetectionStrategy.QUANTILES: DetectionStrategy.IQR,
    DetectionStrategy.TAIL_AWARE: DetectionStrategy.IQR,
    DetectionStrategy.WINDOW_COUNT: DetectionStrategy.POISSON,
    DetectionStrategy.SEMANTIC_RULE: DetectionStrategy.POISSON,  # Changed from SKIP to POISSON
}


def normalize_strategy(strategy: DetectionStrategy) -> DetectionStrategy:
    """Convert legacy strategy to core strategy (ivy0205).
    
    All legacy strategies are automatically mapped to their core equivalents.
    This function should be called whenever a strategy is used for detection.
    """
    # Handle string input
    if isinstance(strategy, str):
        try:
            strategy = DetectionStrategy(strategy.upper())
        except ValueError:
            return DetectionStrategy.STANDARD_SPC
    
    return LEGACY_STRATEGY_MAP.get(strategy, strategy)


# Strategy mapping - using core strategies (ivy0203)
KPI_TYPE_TO_STRATEGY = {
    KPIType.CEILING_SATURATED: DetectionStrategy.ABS_THRESHOLD,
    KPIType.FLOOR_SATURATED: DetectionStrategy.ABS_THRESHOLD,
    KPIType.COUNT_TYPE: DetectionStrategy.IQR,           # was RANGE_CONTEXT
    KPIType.RATIO_TYPE: DetectionStrategy.RELATIVE,      # was RELATIVE_CHANGE
    KPIType.PHYSICAL_STABLE: DetectionStrategy.Z_SCORE,  # was ABS_DELTA_TREND
    KPIType.THROUGHPUT_TYPE: DetectionStrategy.IQR,      # was QUANTILES
    KPIType.LATENCY_TYPE: DetectionStrategy.IQR,         # was TAIL_AWARE
    KPIType.LOW_FREQUENCY: DetectionStrategy.POISSON,    # was WINDOW_COUNT
    KPIType.CONFIG_TYPE: DetectionStrategy.ABS_THRESHOLD, # was SKIP (σ=0 constant values)
    KPIType.UNKNOWN: DetectionStrategy.STANDARD_SPC,
}


# Forbidden methods for each type
KPI_TYPE_FORBIDDEN = {
    KPIType.CEILING_SATURATED: ['SPC', '±kσ', 'T-test'],
    KPIType.FLOOR_SATURATED: ['SPC', '±kσ', 'T-test'],
    KPIType.COUNT_TYPE: ['Small sample T-test'],
    KPIType.RATIO_TYPE: ['Normal assumption'],
    KPIType.PHYSICAL_STABLE: ['Single point ±2σ'],
    KPIType.THROUGHPUT_TYPE: ['Mean ±σ'],
    KPIType.LATENCY_TYPE: ['Mean-based methods'],
    KPIType.LOW_FREQUENCY: ['SPC', 'σ-based'],
    KPIType.CONFIG_TYPE: ['All statistical'],
    KPIType.UNKNOWN: [],
}


@dataclass
class KPIDefinitionInfo:
    """KPI definition from XML"""
    id: str
    name: str
    level: str = ''
    formula: str = ''
    units: str = ''
    desc: str = ''
    related_counters: List[str] = field(default_factory=list)


@dataclass
class CounterDefinitionInfo:
    """Counter definition from XML"""
    name: str
    netact: str = ''
    units: str = ''
    time_aggregation: str = ''
    desc: str = ''


@dataclass
class KPIClassificationInfo:
    """Classification information for a KPI/Counter"""
    code: str
    name: str
    code_type: str  # 'KPI' or 'Counter'
    units: str
    kpi_type: KPIType
    detection_strategy: DetectionStrategy
    variance_type: str  # ceiling/floor/constant/stable/moderate/high_variance
    xml_found: bool
    level: str
    formula: str
    related_counters: List[str]
    thresholds: Dict[str, Any]
    forbidden_methods: List[str]
    classification_reason: str
    
    def to_dict(self) -> Dict:
        """Convert to dictionary"""
        return {
            'code': self.code,
            'name': self.name,
            'code_type': self.code_type,
            'units': self.units,
            'kpi_type': self.kpi_type.value,
            'detection_strategy': self.detection_strategy.value,
            'variance_type': self.variance_type,
            'xml_found': self.xml_found,
            'level': self.level,
            'formula': self.formula,
            'related_counters': self.related_counters,
            'thresholds': self.thresholds,
            'forbidden_methods': self.forbidden_methods,
            'classification_reason': self.classification_reason,
        }


class KPITypeClassifier:
    """KPI Type Classification based on XML definitions and historical data"""
    
    def __init__(self, lib_dir: Path = None, config_file: Path = None):
        """Initialize classifier (ivy0205 simplified)
        
        Args:
            lib_dir: Directory containing XML definition files
            config_file: Path to unified config Excel file (kpi_config_unified.xlsx)
        """
        self.lib_dir = lib_dir or Path(__file__).resolve().parent / 'assets' / 'kpi_xml'
        
        # Single source of truth: kpi_config_unified.xlsx
        self.config_file = config_file or self.lib_dir / 'kpi_config_unified.xlsx'
        
        # Legacy file paths (for backward compatibility)
        self.classification_file = self.lib_dir / 'kpi_classification_v2.xlsx'
        
        self.kpi_defs: Dict[str, KPIDefinitionInfo] = {}
        self.counter_defs: Dict[str, CounterDefinitionInfo] = {}
        self.classifications: Dict[str, KPIClassificationInfo] = {}
        
        # Reverse lookup: counter -> list of KPIs that use this counter
        self.counter_to_kpis: Dict[str, List[str]] = {}
        
        # Custom rules storage (loaded from unified file)
        self.custom_rules: List[Dict] = []
        self.custom_code_rules: Dict[str, DetectionStrategy] = {}  # code -> strategy
        self.custom_pattern_rules: List[Tuple[str, DetectionStrategy]] = []  # (pattern, strategy)
        self.custom_type_rules: Dict[str, DetectionStrategy] = {}  # kpi_type -> strategy
        
        self._loaded = False
    
    def load(self) -> bool:
        """Load classification data from unified config file"""
        if self._loaded:
            return True
        
        # Try unified file first, then fall back to legacy
        config_path = self.config_file if self.config_file.exists() else self.classification_file
        
        if config_path.exists():
            try:
                self._load_from_unified_file(config_path)
                self._loaded = True
                # Also load XML for SCOUT_NR lookup
                self._load_from_xml()
                # Build reverse lookup map
                self._build_counter_to_kpis_map()
                print(f"KPI Type Classifier: Loaded {len(self.classifications)} classifications from {config_path.name}")
                print(f"KPI Type Classifier: Built counter→KPI map with {len(self.counter_to_kpis)} counters")
                if self.custom_code_rules:
                    print(f"KPI Type Classifier: Loaded {len(self.custom_code_rules)} custom overrides")
                return True
            except Exception as e:
                print(f"Warning: Failed to load config file - {e}")
        
        # Fall back to parsing XML files
        self._load_from_xml()
        # Build reverse lookup map
        self._build_counter_to_kpis_map()
        self._loaded = True
        return True
    
    def _load_from_unified_file(self, config_path: Path):
        """Load from unified config file (contains both classification and override rules)"""
        # Try multiple sheet names for compatibility
        sheet_names_to_try = ['1_All_Codes_Classification', 'Sheet1', 0]
        df = None
        for sheet_name in sheet_names_to_try:
            try:
                df = pd.read_excel(config_path, sheet_name=sheet_name)
                break
            except Exception:
                continue
        
        if df is None:
            raise ValueError(f"Could not load any sheet from {config_path}")
        
        # Load classifications and custom rules in one pass
        for _, row in df.iterrows():
            code = str(row.get('Code', '')).strip()
            if not code:
                continue
            
            # Load classification info
            try:
                kpi_type = KPIType(str(row.get('KPI_Type', 'unknown')).lower())
            except ValueError:
                kpi_type = KPIType.UNKNOWN
            
            try:
                detection_strategy = DetectionStrategy(str(row.get('Detection_Strategy', 'STANDARD_SPC')).upper())
            except ValueError:
                detection_strategy = DetectionStrategy.STANDARD_SPC
            
            # Parse related counters
            related_counters = []
            rc_str = row.get('Related_Counters', '')
            if pd.notna(rc_str) and rc_str:
                related_counters = [c.strip() for c in str(rc_str).split(',') if c.strip()]
            
            # Parse thresholds
            thresholds = {}
            threshold_str = row.get('Thresholds', '')
            if pd.notna(threshold_str) and threshold_str:
                thresholds = self._parse_threshold_string(str(threshold_str))
            
            # Parse forbidden methods
            forbidden = []
            forbidden_str = row.get('Forbidden_Methods', '')
            if pd.notna(forbidden_str) and forbidden_str:
                forbidden = [m.strip() for m in str(forbidden_str).split(',') if m.strip()]
            
            # Create classification info
            self.classifications[code] = KPIClassificationInfo(
                code=code,
                name=str(row.get('Name', '')) if pd.notna(row.get('Name')) else '',
                code_type=str(row.get('Type', 'Other')) if pd.notna(row.get('Type')) else 'Other',
                units=str(row.get('Units', '')) if pd.notna(row.get('Units')) else '',
                kpi_type=kpi_type,
                detection_strategy=detection_strategy,
                variance_type=str(row.get('Variance_Type', '')) if pd.notna(row.get('Variance_Type')) else '',
                xml_found=(str(row.get('XML_Found', '')).strip() == '✓'),
                level=str(row.get('Level', '')) if pd.notna(row.get('Level')) else '',
                formula=str(row.get('Formula', '')) if pd.notna(row.get('Formula')) else '',
                related_counters=related_counters,
                thresholds=thresholds,
                forbidden_methods=forbidden,
                classification_reason=str(row.get('Classification_Reason', '')) if pd.notna(row.get('Classification_Reason')) else '',
            )
            
            # Load custom override (if Enabled=True)
            if 'Override_Strategy' in df.columns and 'Enabled' in df.columns:
                override = str(row.get('Override_Strategy', '')).strip().upper()
                enabled_val = row.get('Enabled', False)
                enabled = (enabled_val is True) or (str(enabled_val).strip().upper() == 'TRUE')
                
                if enabled and override and override not in ('', 'NAN', '(使用默认)'):
                    try:
                        strategy = DetectionStrategy(override)
                        self.custom_code_rules[code] = strategy
                    except ValueError:
                        pass
    
    def _parse_threshold_string(self, s: str) -> Dict[str, Any]:
        """Parse threshold string like 'Normal=100%, Acceptable≥99.9%'"""
        thresholds = {}
        # Simple parsing - can be enhanced
        if 'Normal' in s:
            thresholds['normal'] = s
        return thresholds
    
    def _load_from_xml(self):
        """Load from XML definition files"""
        # Load KPI definitions
        for xml_file in ['PMKPI_SBTS26R2_5G.xml', 'JUMP_KPIs.xml']:
            xml_path = self.lib_dir / xml_file
            if xml_path.exists():
                self._parse_kpi_xml(xml_path)
        
        # Load Counter definitions
        counter_xml = self.lib_dir / 'PMPegs_SBTS26R2_5G.xml'
        if counter_xml.exists():
            self._parse_counter_xml(counter_xml)
    
    def _parse_kpi_xml(self, xml_path: Path):
        """Parse KPI XML file"""
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
        except ET.ParseError:
            with open(xml_path, 'r', encoding='utf-8') as f:
                content = f.read()
            if not content.strip().startswith('<?xml'):
                content = '<?xml version="1.0"?>\n' + content
            root = ET.fromstring(content)
        
        for kpi_elem in root.findall('.//KPI'):
            kpi_id = kpi_elem.get('id', '')
            if not kpi_id:
                continue
            
            formula = kpi_elem.get('formula', '')
            related = self._extract_counters_from_formula(formula)
            
            self.kpi_defs[kpi_id] = KPIDefinitionInfo(
                id=kpi_id,
                name=kpi_elem.get('name', ''),
                level=kpi_elem.get('level', ''),
                formula=formula,
                units=kpi_elem.get('units', ''),
                desc=kpi_elem.get('desc', ''),
                related_counters=related,
            )
    
    def _parse_counter_xml(self, xml_path: Path):
        """Parse Counter XML file"""
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
        except ET.ParseError:
            return
        
        for measurement in root.findall('.//Measurement'):
            for peg in measurement.findall('peg'):
                counter_name = peg.get('name', '')
                if not counter_name:
                    continue
                
                self.counter_defs[counter_name] = CounterDefinitionInfo(
                    name=counter_name,
                    netact=peg.get('NetAct', ''),
                    units=peg.get('units', ''),
                    time_aggregation=peg.get('timeAggregation', ''),
                    desc=peg.get('desc', ''),
                )
    
    def _extract_counters_from_formula(self, formula: str) -> List[str]:
        """Extract counter references from formula"""
        if not formula:
            return []
        pattern = r'M\d+C\d+'
        return list(set(re.findall(pattern, formula)))
    
    def _normalize_code(self, code: str) -> str:
        """Normalize code for lookup (handle SCOUNT_NR_ prefix)"""
        if code.startswith('SCOUNT_'):
            return code[7:]
        return code
    
    def _normalize_scout_to_nr(self, code: str) -> Optional[str]:
        """Convert SCOUT_NR_xy to NR_xy format for XML lookup
        
        Args:
            code: Code like 'SCOUT_NR_1234a' or 'NR_1234a'
            
        Returns:
            NR_xy format string, or None if not a valid pattern
        """
        import re
        # Match SCOUT_NR_xxxx[a-z] pattern
        match = re.match(r'^SCOUT_(NR_\d{2,4}[a-zA-Z])$', code)
        if match:
            return match.group(1)
        # Already NR_ format
        match = re.match(r'^(NR_\d{2,4}[a-zA-Z])$', code)
        if match:
            return match.group(1)
        return None
    
    def _build_counter_to_kpis_map(self):
        """Build reverse lookup map: counter -> list of KPIs"""
        self.counter_to_kpis = {}
        
        # Build from classifications
        for code, cls in self.classifications.items():
            if cls.related_counters:
                for counter in cls.related_counters:
                    if counter not in self.counter_to_kpis:
                        self.counter_to_kpis[counter] = []
                    if code not in self.counter_to_kpis[counter]:
                        self.counter_to_kpis[counter].append(code)
        
        # Build from KPI definitions
        for kpi_id, kpi_def in self.kpi_defs.items():
            if kpi_def.related_counters:
                for counter in kpi_def.related_counters:
                    if counter not in self.counter_to_kpis:
                        self.counter_to_kpis[counter] = []
                    if kpi_id not in self.counter_to_kpis[counter]:
                        self.counter_to_kpis[counter].append(kpi_id)

    def get_classification(self, code: str) -> Optional[KPIClassificationInfo]:
        """Get classification for a code"""
        self.load()
        
        # Direct lookup
        if code in self.classifications:
            return self.classifications[code]
        
        # Normalized lookup
        norm_code = self._normalize_code(code)
        if norm_code in self.classifications:
            return self.classifications[norm_code]
        
        return None
    
    def _get_custom_strategy(self, code: str, kpi_type: KPIType) -> Optional[DetectionStrategy]:
        """Check if there's a custom strategy override for this code
        
        Priority: codes > pattern > kpi_type
        
        Returns:
            Custom DetectionStrategy if found, None otherwise
        """
        # 1. Check code-based rules (highest priority)
        if code in self.custom_code_rules:
            return self.custom_code_rules[code]
        
        # 2. Check pattern-based rules
        for pattern, strategy in self.custom_pattern_rules:
            if re.match(pattern, code):
                return strategy
        
        # 3. Check kpi_type-based rules
        kpi_type_str = kpi_type.value if kpi_type else 'unknown'
        if kpi_type_str in self.custom_type_rules:
            return self.custom_type_rules[kpi_type_str]
        
        return None
    
    def get_kpi_type(self, code: str) -> KPIType:
        """Get KPI type for a code"""
        cls = self.get_classification(code)
        return cls.kpi_type if cls else KPIType.UNKNOWN
    
    def get_detection_strategy(self, code: str) -> DetectionStrategy:
        """Get detection strategy for a code
        
        Custom rules override default strategy with priority:
        codes > pattern > kpi_type > default
        """
        cls = self.get_classification(code)
        kpi_type = cls.kpi_type if cls else KPIType.UNKNOWN
        default_strategy = cls.detection_strategy if cls else DetectionStrategy.STANDARD_SPC
        
        # Check for custom override
        custom_strategy = self._get_custom_strategy(code, kpi_type)
        if custom_strategy:
            return custom_strategy
        
        return default_strategy
    
    def is_custom_rule_applied(self, code: str) -> bool:
        """Check if a custom rule is applied to this code"""
        cls = self.get_classification(code)
        kpi_type = cls.kpi_type if cls else KPIType.UNKNOWN
        return self._get_custom_strategy(code, kpi_type) is not None
    
    def get_related_counters(self, code: str) -> List[str]:
        """Get related counters for a KPI
        
        For SCOUT_NR_xy codes, also try NR_xy lookup in XML definitions
        """
        self.load()
        
        # First try from classification
        cls = self.get_classification(code)
        if cls and cls.related_counters:
            return cls.related_counters
        
        # For SCOUT_NR_ codes, try NR_ lookup in kpi_defs
        nr_code = self._normalize_scout_to_nr(code)
        if nr_code and nr_code in self.kpi_defs:
            kpi_def = self.kpi_defs[nr_code]
            if kpi_def.related_counters:
                return kpi_def.related_counters
        
        return []
    
    def get_formula(self, code: str) -> str:
        """Get formula for a KPI
        
        For SCOUT_NR_xy codes, also try NR_xy lookup in XML definitions
        """
        self.load()
        
        # First try from classification
        cls = self.get_classification(code)
        if cls and cls.formula:
            return cls.formula
        
        # For SCOUT_NR_ codes, try NR_ lookup in kpi_defs
        nr_code = self._normalize_scout_to_nr(code)
        if nr_code and nr_code in self.kpi_defs:
            kpi_def = self.kpi_defs[nr_code]
            if kpi_def.formula:
                return kpi_def.formula
        
        return ''
    
    def get_related_kpis(self, counter_code: str) -> List[str]:
        """Get list of KPIs that use this counter
        
        Args:
            counter_code: Counter code like 'M55124C00006'
            
        Returns:
            List of KPI codes that reference this counter in their formula
        """
        self.load()
        return self.counter_to_kpis.get(counter_code, [])
    
    def get_units(self, code: str) -> str:
        """Get units for a code"""
        cls = self.get_classification(code)
        return cls.units if cls else ''
    
    def get_kpi_definition(self, code: str) -> Optional[KPIDefinitionInfo]:
        """Get KPI definition"""
        self.load()
        norm_code = self._normalize_code(code)
        return self.kpi_defs.get(code) or self.kpi_defs.get(norm_code)
    
    def get_counter_definition(self, code: str) -> Optional[CounterDefinitionInfo]:
        """Get Counter definition"""
        self.load()
        return self.counter_defs.get(code)

    def search_code_in_xml(self, code: str) -> Optional[Dict[str, str]]:
        """Search for a code in XML definitions (KPI + Counter XMLs)
        
        Returns:
            Dict with code info if found, None otherwise
        """
        self.load()
        
        # Try direct lookup in kpi_defs
        kpi_def = self.kpi_defs.get(code)
        if not kpi_def:
            # Try normalized code (SCOUT_NR_ -> NR_)
            norm_code = self._normalize_scout_to_nr(code)
            if norm_code:
                kpi_def = self.kpi_defs.get(norm_code)
        
        if kpi_def:
            return {
                'code': code,
                'name': kpi_def.name,
                'type': 'KPI',
                'units': kpi_def.units,
                'level': kpi_def.level,
                'formula': kpi_def.formula,
                'xml_found': True,
                'related_counters': ','.join(kpi_def.related_counters) if kpi_def.related_counters else ''
            }
        
        # Try counter_defs
        counter_def = self.counter_defs.get(code)
        if counter_def:
            return {
                'code': code,
                'name': counter_def.name,
                'type': 'Counter',
                'units': counter_def.units,
                'level': '',
                'formula': '',
                'xml_found': True,
                'related_counters': ''
            }
        
        return None

    def register_new_codes(self, codes: List[str], default_strategy: str = 'IQR') -> List[str]:
        """Search for unknown codes in XML and append to kpi_config_unified.xlsx
        
        Args:
            codes: List of codes to check
            default_strategy: Default detection strategy for new codes
            
        Returns:
            List of newly added codes
        """
        self.load()
        
        new_codes = []
        new_rows = []
        
        for code in codes:
            # Skip if already in classifications
            if code in self.classifications:
                continue
            norm_code = self._normalize_code(code)
            if norm_code in self.classifications:
                continue
            
            # Search in XML
            xml_info = self.search_code_in_xml(code)
            
            if xml_info:
                new_row = {
                    'Code': code,
                    'Name': xml_info.get('name', ''),
                    'Type': xml_info.get('type', 'Other'),
                    'Units': xml_info.get('units', ''),
                    'KPI_Type': 'unknown',
                    'Detection_Strategy': default_strategy,
                    'Variance_Type': '',
                    'XML_Found': '✓',
                    'Level': xml_info.get('level', ''),
                    'Formula': xml_info.get('formula', ''),
                    'Related_Counters': xml_info.get('related_counters', ''),
                    'Thresholds': '',
                    'Forbidden_Methods': '',
                    'Classification_Reason': 'Auto-added: found in XML',
                    'Override_Strategy': '',
                    'Enabled': False,
                    'Constraint': '',
                }
            else:
                # Not found in XML either, still add with minimal info
                from .utils import classify_code
                code_type = classify_code(code)
                new_row = {
                    'Code': code,
                    'Name': '',
                    'Type': code_type,
                    'Units': '',
                    'KPI_Type': 'unknown',
                    'Detection_Strategy': default_strategy,
                    'Variance_Type': '',
                    'XML_Found': '',
                    'Level': '',
                    'Formula': '',
                    'Related_Counters': '',
                    'Thresholds': '',
                    'Forbidden_Methods': '',
                    'Classification_Reason': 'Auto-added: not found in XML',
                    'Override_Strategy': '',
                    'Enabled': False,
                    'Constraint': '',
                }
            
            new_rows.append(new_row)
            new_codes.append(code)
            
            # Register in memory
            try:
                det_strategy = DetectionStrategy(default_strategy.upper())
            except ValueError:
                det_strategy = DetectionStrategy.IQR
            
            self.classifications[code] = KPIClassificationInfo(
                code=code,
                name=new_row['Name'],
                code_type=new_row['Type'],
                units=new_row['Units'],
                kpi_type=KPIType.UNKNOWN,
                detection_strategy=det_strategy,
                variance_type='',
                xml_found=bool(xml_info),
                level=new_row['Level'],
                formula=new_row['Formula'],
                related_counters=new_row['Related_Counters'].split(',') if new_row['Related_Counters'] else [],
                thresholds={},
                forbidden_methods=[],
                classification_reason=new_row['Classification_Reason'],
            )
        
        # Append to xlsx file if there are new codes
        if new_rows and self.config_file.exists():
            try:
                self._append_to_unified_xlsx(new_rows)
                print(f"KPI Type Classifier: Appended {len(new_rows)} new codes to {self.config_file.name}")
                for c in new_codes:
                    xml_status = 'XML found' if self.search_code_in_xml(c) else 'not in XML'
                    print(f"  + {c} ({xml_status}, strategy={default_strategy})")
            except Exception as e:
                print(f"Warning: Failed to append new codes to xlsx - {e}")
        
        return new_codes

    def _append_to_unified_xlsx(self, new_rows: List[Dict]):
        """Append new rows to kpi_config_unified.xlsx without changing existing content/format
        
        Uses openpyxl to preserve formatting
        """
        import openpyxl
        
        wb = openpyxl.load_workbook(self.config_file)
        
        # Find the right sheet
        sheet_names_to_try = ['1_All_Codes_Classification', 'Sheet1']
        ws = None
        for name in sheet_names_to_try:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active
        
        # Read header row to get column mapping
        header_map = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header:
                header_map[str(header).strip()] = col_idx
        
        # Append new rows
        next_row = ws.max_row + 1
        for row_data in new_rows:
            for key, value in row_data.items():
                if key in header_map:
                    ws.cell(row=next_row, column=header_map[key], value=value)
            next_row += 1
        
        wb.save(self.config_file)
        wb.close()


class TypeBasedAnomalyDetector:
    """Type-based anomaly detection that applies different strategies per KPI type"""
    
    def __init__(self, classifier: KPITypeClassifier = None):
        """Initialize detector
        
        Args:
            classifier: KPI type classifier instance
        """
        self.classifier = classifier or KPITypeClassifier()
    
    def detect_anomaly(self, code: str, current_values: List[float],
                       hist_mean: float, hist_std: float,
                       history_records: List[Dict] = None) -> Dict[str, Any]:
        """Detect anomaly using type-appropriate strategy
        
        Args:
            code: KPI/Counter code
            current_values: Current data values
            hist_mean: Historical mean
            hist_std: Historical standard deviation
            history_records: Historical records for trend analysis
            
        Returns:
            Dictionary with detection results
        """
        if not current_values:
            return self._create_result('NoData', 'No current data')
        
        current_mean = np.mean(current_values)
        kpi_type = self.classifier.get_kpi_type(code)
        strategy = self.classifier.get_detection_strategy(code)
        
        # Apply type-specific detection
        if strategy == DetectionStrategy.ABS_THRESHOLD:
            return self._detect_abs_threshold(code, current_values, current_mean, hist_mean, kpi_type)
        elif strategy == DetectionStrategy.RANGE_CONTEXT:
            return self._detect_range_context(code, current_values, current_mean, hist_mean, hist_std)
        elif strategy == DetectionStrategy.RELATIVE_CHANGE:
            return self._detect_relative_change(code, current_values, current_mean, hist_mean, hist_std)
        elif strategy == DetectionStrategy.ABS_DELTA_TREND:
            return self._detect_abs_delta_trend(code, current_values, current_mean, hist_mean, hist_std, history_records)
        elif strategy == DetectionStrategy.QUANTILES:
            return self._detect_quantiles(code, current_values, hist_mean, hist_std, history_records)
        elif strategy == DetectionStrategy.TAIL_AWARE:
            return self._detect_tail_aware(code, current_values, hist_mean, hist_std, history_records)
        elif strategy == DetectionStrategy.WINDOW_COUNT:
            return self._detect_window_count(code, current_values, hist_mean)
        elif strategy == DetectionStrategy.SEMANTIC_RULE:
            return self._detect_semantic_rule(code, current_values, current_mean, hist_mean)
        else:
            return self._detect_standard_spc(code, current_values, current_mean, hist_mean, hist_std)
    
    def _create_result(self, severity: str, reason: str, **kwargs) -> Dict[str, Any]:
        """Create standard result dictionary"""
        result = {
            'severity': severity,
            'reasons': [reason] if reason else [],
            'kpi_type': kwargs.get('kpi_type', KPIType.UNKNOWN).value if isinstance(kwargs.get('kpi_type'), KPIType) else kwargs.get('kpi_type', 'unknown'),
            'detection_strategy': kwargs.get('strategy', DetectionStrategy.STANDARD_SPC).value if isinstance(kwargs.get('strategy'), DetectionStrategy) else kwargs.get('strategy', 'STANDARD_SPC'),
            'anomaly_count': kwargs.get('anomaly_count', 0),
            'anomaly_indices': kwargs.get('anomaly_indices', []),
            'level_shift_detected': kwargs.get('level_shift_detected', False),
            'trend_direction': kwargs.get('trend_direction', None),
        }
        result.update({k: v for k, v in kwargs.items() if k not in result})
        return result
    
    def _detect_abs_threshold(self, code: str, current_values: List[float],
                               current_mean: float, hist_mean: float,
                               kpi_type: KPIType) -> Dict[str, Any]:
        """ABS_THRESHOLD detection for ceiling/floor saturated KPIs"""
        cls = self.classifier.get_classification(code)
        
        if kpi_type == KPIType.CEILING_SATURATED:
            # Ceiling KPI (history≈100%): success rates
            if current_mean == 100:
                severity = 'Normal'
                reason = 'Ceiling KPI: current=100% (Stable)'
            elif current_mean >= 99.9:
                severity = 'Normal'
                reason = f'Ceiling KPI: current={current_mean:.2f}% ≥99.9% (Acceptable)'
            elif current_mean >= 99:
                severity = 'Suspicious'
                reason = f'Ceiling KPI: current={current_mean:.2f}% (99%~99.9%, Minor drop)'
            else:
                severity = 'Regression'
                reason = f'Ceiling KPI: current={current_mean:.2f}% <99% (Significant drop)'
            
            # Check individual values
            min_val = min(current_values)
            anomaly_indices = [i for i, v in enumerate(current_values) if v < 99]
            if min_val < 99 and severity == 'Normal':
                severity = 'Suspicious'
                reason += f'; min={min_val:.2f}%'
            
        elif kpi_type == KPIType.FLOOR_SATURATED:
            # Floor KPI (history≈0): failure/error counts
            if current_mean == 0:
                severity = 'Normal'
                reason = 'Floor KPI: current=0 (Stable)'
            elif current_mean <= 0.5:
                severity = 'Normal'
                reason = f'Floor KPI: current={current_mean:.3f} ≤0.5 (Minor fluctuation)'
            elif current_mean <= 2:
                severity = 'Suspicious'
                reason = f'Floor KPI: current={current_mean:.3f} (0.5~2, Warning)'
            else:
                severity = 'Regression'
                reason = f'Floor KPI: current={current_mean:.3f} >2 (Significant increase)'
            
            # Check individual values
            max_val = max(current_values)
            anomaly_indices = [i for i, v in enumerate(current_values) if v > 2]
            if max_val > 5 and severity != 'Regression':
                severity = 'Regression'
                reason += f'; max={max_val:.3f}'
        else:
            # Generic constant KPI
            delta = current_mean - hist_mean
            if delta == 0:
                severity = 'Normal'
                reason = 'Constant KPI: current=history (Stable)'
            elif abs(hist_mean) >= 10:
                rel_diff = abs(delta / hist_mean) * 100
                if rel_diff <= 0.5:
                    severity = 'Normal'
                    reason = f'Constant KPI: Δ={delta:.2f} ({rel_diff:.2f}%)'
                elif rel_diff <= 2:
                    severity = 'Suspicious'
                    reason = f'Constant KPI: Δ={delta:.2f} ({rel_diff:.2f}%)'
                else:
                    severity = 'Regression'
                    reason = f'Constant KPI: Δ={delta:.2f} ({rel_diff:.2f}%)'
            else:
                if abs(delta) <= 0.1:
                    severity = 'Normal'
                    reason = f'Constant KPI: Δ={delta:.4f}'
                elif abs(delta) <= 1:
                    severity = 'Suspicious'
                    reason = f'Constant KPI: Δ={delta:.4f}'
                else:
                    severity = 'Regression'
                    reason = f'Constant KPI: Δ={delta:.4f}'
            anomaly_indices = []
        
        return self._create_result(severity, reason, kpi_type=kpi_type, 
                                   strategy=DetectionStrategy.ABS_THRESHOLD,
                                   anomaly_count=len(anomaly_indices),
                                   anomaly_indices=anomaly_indices)
    
    def _detect_range_context(self, code: str, current_values: List[float],
                               current_mean: float, hist_mean: float,
                               hist_std: float) -> Dict[str, Any]:
        """RANGE_CONTEXT detection for count type KPIs"""
        kpi_type = KPIType.COUNT_TYPE
        
        if hist_std == 0:
            # All history identical
            if current_mean == hist_mean:
                return self._create_result('Normal', 'Count KPI: current=history', 
                                          kpi_type=kpi_type, strategy=DetectionStrategy.RANGE_CONTEXT)
            else:
                delta_pct = abs(current_mean - hist_mean) / (hist_mean + 0.01) * 100
                if delta_pct <= 10:
                    return self._create_result('Normal', f'Count KPI: Δ={delta_pct:.1f}%',
                                              kpi_type=kpi_type, strategy=DetectionStrategy.RANGE_CONTEXT)
                elif delta_pct <= 30:
                    return self._create_result('Suspicious', f'Count KPI: Δ={delta_pct:.1f}%',
                                              kpi_type=kpi_type, strategy=DetectionStrategy.RANGE_CONTEXT)
                else:
                    return self._create_result('Regression', f'Count KPI: Δ={delta_pct:.1f}%',
                                              kpi_type=kpi_type, strategy=DetectionStrategy.RANGE_CONTEXT)
        
        # Range check: μ ± 2σ
        upper = hist_mean + 2 * hist_std
        lower = hist_mean - 2 * hist_std
        
        anomaly_indices = [i for i, v in enumerate(current_values) if v > upper or v < lower]
        anomaly_ratio = len(anomaly_indices) / len(current_values)
        
        if lower <= current_mean <= upper:
            if anomaly_ratio <= 0.1:
                severity = 'Normal'
                reason = f'Count KPI: mean in [μ±2σ] range'
            else:
                severity = 'Suspicious'
                reason = f'Count KPI: mean in range but {anomaly_ratio:.0%} points outside'
        elif current_mean > hist_mean + 3 * hist_std or current_mean < hist_mean - 3 * hist_std:
            severity = 'Regression'
            reason = f'Count KPI: mean exceeds ±3σ'
        else:
            severity = 'Suspicious'
            reason = f'Count KPI: mean exceeds ±2σ'
        
        return self._create_result(severity, reason, kpi_type=kpi_type,
                                   strategy=DetectionStrategy.RANGE_CONTEXT,
                                   anomaly_count=len(anomaly_indices),
                                   anomaly_indices=anomaly_indices)
    
    def _detect_relative_change(self, code: str, current_values: List[float],
                                 current_mean: float, hist_mean: float,
                                 hist_std: float) -> Dict[str, Any]:
        """RELATIVE_CHANGE detection for ratio type KPIs"""
        kpi_type = KPIType.RATIO_TYPE
        
        # Calculate relative change
        if hist_mean == 0:
            rel_change = float('inf') if current_mean != 0 else 0
        else:
            rel_change = (current_mean - hist_mean) / hist_mean * 100
        
        # Thresholds for ratio KPIs (like BLER, drop rate)
        if abs(rel_change) <= 5:
            severity = 'Normal'
            reason = f'Ratio KPI: Δ%={rel_change:.1f}% (≤5%)'
        elif abs(rel_change) <= 15:
            severity = 'Suspicious'
            reason = f'Ratio KPI: Δ%={rel_change:.1f}% (5%~15%)'
        else:
            severity = 'Regression'
            reason = f'Ratio KPI: Δ%={rel_change:.1f}% (>15%)'
        
        # Find anomaly points using relative threshold
        threshold = hist_mean * 0.2 if hist_mean != 0 else hist_std * 2
        anomaly_indices = [i for i, v in enumerate(current_values) 
                          if abs(v - hist_mean) > threshold]
        
        return self._create_result(severity, reason, kpi_type=kpi_type,
                                   strategy=DetectionStrategy.RELATIVE_CHANGE,
                                   anomaly_count=len(anomaly_indices),
                                   anomaly_indices=anomaly_indices,
                                   relative_change=rel_change)
    
    def _detect_abs_delta_trend(self, code: str, current_values: List[float],
                                 current_mean: float, hist_mean: float, hist_std: float,
                                 history_records: List[Dict] = None) -> Dict[str, Any]:
        """ABS_DELTA_TREND detection for physical stable KPIs (dB, dBm)"""
        kpi_type = KPIType.PHYSICAL_STABLE
        
        # Absolute delta threshold (for dB values)
        delta = current_mean - hist_mean
        
        # Level shift detection
        level_shift = False
        trend_direction = None
        
        if history_records:
            hist_means = self._extract_hist_means(code, history_records)
            if len(hist_means) >= 3:
                level_shift, trend_direction = self._detect_level_shift(hist_means, current_mean)
        
        # Thresholds for physical measurements
        if abs(delta) <= 0.2:
            severity = 'Normal'
            reason = f'Physical KPI: Δ={delta:.3f} dB (≤0.2)'
        elif abs(delta) <= 0.5:
            severity = 'Suspicious'
            reason = f'Physical KPI: Δ={delta:.3f} dB (0.2~0.5)'
        elif abs(delta) <= 1.0:
            severity = 'Suspicious' if not level_shift else 'Regression'
            reason = f'Physical KPI: Δ={delta:.3f} dB (0.5~1.0)'
        else:
            severity = 'Regression'
            reason = f'Physical KPI: Δ={delta:.3f} dB (>1.0)'
        
        if level_shift:
            severity = 'Regression' if severity != 'Regression' else severity
            reason += f'; Level shift detected ({trend_direction})'
        
        # Anomaly points
        threshold = 0.5  # dB threshold for individual points
        anomaly_indices = [i for i, v in enumerate(current_values)
                          if abs(v - hist_mean) > threshold]
        
        return self._create_result(severity, reason, kpi_type=kpi_type,
                                   strategy=DetectionStrategy.ABS_DELTA_TREND,
                                   anomaly_count=len(anomaly_indices),
                                   anomaly_indices=anomaly_indices,
                                   level_shift_detected=level_shift,
                                   trend_direction=trend_direction)
    
    def _detect_quantiles(self, code: str, current_values: List[float],
                          hist_mean: float, hist_std: float,
                          history_records: List[Dict] = None) -> Dict[str, Any]:
        """QUANTILES detection for throughput type KPIs"""
        kpi_type = KPIType.THROUGHPUT_TYPE
        
        current_median = np.median(current_values)
        current_p95 = np.percentile(current_values, 95) if len(current_values) >= 5 else max(current_values)
        
        # Get historical percentiles if available
        hist_p50 = hist_mean  # Approximate
        
        # Relative median change
        if hist_p50 == 0:
            median_change = 0 if current_median == 0 else float('inf')
        else:
            median_change = (current_median - hist_p50) / hist_p50 * 100
        
        if abs(median_change) <= 10:
            severity = 'Normal'
            reason = f'Throughput KPI: P50 Δ%={median_change:.1f}%'
        elif abs(median_change) <= 25:
            severity = 'Suspicious'
            reason = f'Throughput KPI: P50 Δ%={median_change:.1f}%'
        else:
            severity = 'Regression'
            reason = f'Throughput KPI: P50 Δ%={median_change:.1f}%'
        
        return self._create_result(severity, reason, kpi_type=kpi_type,
                                   strategy=DetectionStrategy.QUANTILES,
                                   current_median=current_median,
                                   current_p95=current_p95)
    
    def _detect_tail_aware(self, code: str, current_values: List[float],
                           hist_mean: float, hist_std: float,
                           history_records: List[Dict] = None) -> Dict[str, Any]:
        """TAIL_AWARE detection for latency type KPIs"""
        kpi_type = KPIType.LATENCY_TYPE
        
        current_p95 = np.percentile(current_values, 95) if len(current_values) >= 5 else max(current_values)
        current_p99 = np.percentile(current_values, 99) if len(current_values) >= 10 else max(current_values)
        
        # Use mean + 2*std as reference P95
        hist_p95_approx = hist_mean + 2 * hist_std if hist_std > 0 else hist_mean * 1.2
        
        if hist_p95_approx == 0:
            p95_change = 0 if current_p95 == 0 else float('inf')
        else:
            p95_change = (current_p95 - hist_p95_approx) / hist_p95_approx * 100
        
        if p95_change <= 10:
            severity = 'Normal'
            reason = f'Latency KPI: P95 Δ%={p95_change:.1f}%'
        elif p95_change <= 30:
            severity = 'Suspicious'
            reason = f'Latency KPI: P95 Δ%={p95_change:.1f}%'
        else:
            severity = 'Regression'
            reason = f'Latency KPI: P95 Δ%={p95_change:.1f}%'
        
        # Check for tail growth
        anomaly_indices = [i for i, v in enumerate(current_values)
                          if v > hist_mean + 3 * hist_std]
        
        return self._create_result(severity, reason, kpi_type=kpi_type,
                                   strategy=DetectionStrategy.TAIL_AWARE,
                                   anomaly_count=len(anomaly_indices),
                                   anomaly_indices=anomaly_indices,
                                   current_p95=current_p95,
                                   current_p99=current_p99)
    
    def _detect_window_count(self, code: str, current_values: List[float],
                              hist_mean: float) -> Dict[str, Any]:
        """WINDOW_COUNT detection for low frequency events"""
        kpi_type = KPIType.LOW_FREQUENCY
        
        current_sum = sum(current_values)
        current_mean = np.mean(current_values)
        
        # For rare events (history ≈ 0)
        if hist_mean <= 0.1:
            if current_sum == 0:
                severity = 'Normal'
                reason = 'Low frequency KPI: no events (Stable)'
            elif current_sum <= 2:
                severity = 'Normal'
                reason = f'Low frequency KPI: {current_sum} events (Acceptable)'
            elif current_sum <= 5:
                severity = 'Suspicious'
                reason = f'Low frequency KPI: {current_sum} events (Warning)'
            else:
                severity = 'Regression'
                reason = f'Low frequency KPI: {current_sum} events (Significant)'
        else:
            # Events expected
            ratio = current_mean / hist_mean if hist_mean > 0 else float('inf')
            if ratio <= 1.5:
                severity = 'Normal'
                reason = f'Low frequency KPI: rate ratio={ratio:.2f}'
            elif ratio <= 3:
                severity = 'Suspicious'
                reason = f'Low frequency KPI: rate ratio={ratio:.2f}'
            else:
                severity = 'Regression'
                reason = f'Low frequency KPI: rate ratio={ratio:.2f}'
        
        # Non-zero points are anomalies for rare events
        anomaly_indices = [i for i, v in enumerate(current_values) if v > 0]
        
        return self._create_result(severity, reason, kpi_type=kpi_type,
                                   strategy=DetectionStrategy.WINDOW_COUNT,
                                   anomaly_count=len(anomaly_indices),
                                   anomaly_indices=anomaly_indices,
                                   event_count=current_sum)
    
    def _detect_semantic_rule(self, code: str, current_values: List[float],
                               current_mean: float, hist_mean: float) -> Dict[str, Any]:
        """SEMANTIC_RULE detection for config type KPIs"""
        kpi_type = KPIType.CONFIG_TYPE
        
        # Config KPIs are typically discrete values
        unique_values = list(set(current_values))
        
        if current_mean == hist_mean:
            severity = 'Normal'
            reason = 'Config KPI: no change (Stable)'
        else:
            # Check if it's a state change
            if len(unique_values) <= 2:
                severity = 'Suspicious'
                reason = f'Config KPI: state change detected'
            else:
                severity = 'Regression'
                reason = f'Config KPI: multiple state changes'
        
        return self._create_result(severity, reason, kpi_type=kpi_type,
                                   strategy=DetectionStrategy.SEMANTIC_RULE,
                                   unique_values=unique_values)
    
    def _detect_standard_spc(self, code: str, current_values: List[float],
                              current_mean: float, hist_mean: float,
                              hist_std: float) -> Dict[str, Any]:
        """Standard SPC detection for unknown types"""
        kpi_type = KPIType.UNKNOWN
        
        # Use near-zero threshold to handle floating-point noise (e.g., 7.1e-15 instead of exact 0)
        hist_std_near_zero = max(abs(hist_mean) * 1e-9, 1e-12)
        if hist_std < hist_std_near_zero:
            # Degenerate case: all historical values essentially identical
            # Use relative deviation to compare (avoids float equality trap)
            rel_dev = abs(current_mean - hist_mean) / (abs(hist_mean) + 1e-10)
            delta_pct = rel_dev * 100
            if delta_pct <= 0.001:  # Less than 0.001% difference → same value
                severity = 'Normal'
                reason = 'Standard SPC: σ≈0, current≈history (Stable)'
            elif delta_pct <= 1:
                severity = 'Normal'
                reason = f'Standard SPC: σ≈0, Δ%={delta_pct:.4f}%'
            elif delta_pct <= 5:
                severity = 'Suspicious'
                reason = f'Standard SPC: σ≈0, Δ%={delta_pct:.2f}%'
            else:
                severity = 'Regression'
                reason = f'Standard SPC: σ≈0, Δ%={delta_pct:.2f}%'
            anomaly_indices = []
        else:
            z_score = (current_mean - hist_mean) / hist_std
            sigma_level = abs(z_score)
            
            if sigma_level <= 2:
                severity = 'Normal'
                reason = f'Standard SPC: σ-level={sigma_level:.2f}'
            elif sigma_level <= 3:
                severity = 'Suspicious'
                reason = f'Standard SPC: σ-level={sigma_level:.2f}'
            else:
                severity = 'Regression'
                reason = f'Standard SPC: σ-level={sigma_level:.2f}'
            
            # Anomaly points
            upper = hist_mean + 2 * hist_std
            lower = hist_mean - 2 * hist_std
            anomaly_indices = [i for i, v in enumerate(current_values)
                              if v > upper or v < lower]
        
        return self._create_result(severity, reason, kpi_type=kpi_type,
                                   strategy=DetectionStrategy.STANDARD_SPC,
                                   anomaly_count=len(anomaly_indices) if 'anomaly_indices' in dir() else 0,
                                   anomaly_indices=anomaly_indices if 'anomaly_indices' in dir() else [])
    
    def _extract_hist_means(self, code: str, history_records: List[Dict]) -> List[float]:
        """Extract historical means for a code"""
        means = []
        for record in history_records:
            kpi_data = record.get('kpi_data', {})
            if code in kpi_data:
                raw_data = kpi_data[code].get('raw_data', '')
                if raw_data:
                    try:
                        values = [float(x) for x in raw_data.split(',') if x.strip()]
                        if values:
                            means.append(np.mean(values))
                    except:
                        pass
        return means
    
    def _detect_level_shift(self, hist_means: List[float], current_mean: float) -> Tuple[bool, Optional[str]]:
        """Detect level shift in time series
        
        Level shift detection using CUSUM-like approach:
        - Check if recent values show persistent deviation from historical baseline
        
        Args:
            hist_means: Historical batch means
            current_mean: Current batch mean
            
        Returns:
            (level_shift_detected, direction)
        """
        if len(hist_means) < 3:
            return False, None
        
        all_means = hist_means + [current_mean]
        n = len(all_means)
        
        # Use first half as baseline
        baseline_n = max(2, n // 2)
        baseline = np.mean(all_means[:baseline_n])
        baseline_std = np.std(all_means[:baseline_n])
        
        if baseline_std == 0:
            baseline_std = abs(baseline) * 0.01 + 0.001  # Prevent division by zero
        
        # Check recent points for persistent shift
        recent_means = all_means[baseline_n:]
        if not recent_means:
            return False, None
        
        # Calculate deviations
        deviations = [(m - baseline) / baseline_std for m in recent_means]
        
        # Level shift if all recent points are on same side and exceed threshold
        if all(d > 1 for d in deviations):
            return True, 'increasing'
        elif all(d < -1 for d in deviations):
            return True, 'decreasing'
        
        # Also check for trend
        if len(recent_means) >= 3:
            diffs = [recent_means[i+1] - recent_means[i] for i in range(len(recent_means)-1)]
            if all(d > 0 for d in diffs):
                return True, 'increasing'
            elif all(d < 0 for d in diffs):
                return True, 'decreasing'
        
        return False, None


# Singleton instance
_classifier_instance: Optional[KPITypeClassifier] = None
_detector_instance: Optional[TypeBasedAnomalyDetector] = None


def get_classifier() -> KPITypeClassifier:
    """Get singleton classifier instance"""
    global _classifier_instance
    if _classifier_instance is None:
        _classifier_instance = KPITypeClassifier()
    return _classifier_instance


def get_type_based_detector() -> TypeBasedAnomalyDetector:
    """Get singleton type-based detector instance"""
    global _detector_instance
    if _detector_instance is None:
        _detector_instance = TypeBasedAnomalyDetector(get_classifier())
    return _detector_instance


# ========== Strategy-Based Detection (ivy0202) ==========

@dataclass
class StrategyConfig:
    """Configuration for strategy-based detection"""
    # IQR method parameters
    iqr_multiplier: float = 1.5       # Standard IQR multiplier for warning
    iqr_strict_multiplier: float = 3.0  # Strict IQR multiplier for regression
    
    # Poisson method parameters
    poisson_alpha: float = 0.05       # Warning threshold p-value
    poisson_alpha_strict: float = 0.01  # Regression threshold p-value
    
    # Relative change parameters
    relative_change_warning: float = 10.0   # Warning: >10% change
    relative_change_critical: float = 25.0  # Regression: >25% change
    
    # Z-score parameters (ivy0203)
    zscore_warning: float = 2.0       # Warning: |z| > 2
    zscore_critical: float = 3.0      # Regression: |z| > 3
    
    # KS-test parameters (ivy0203)
    ks_alpha: float = 0.05            # Warning threshold
    ks_alpha_strict: float = 0.01     # Regression threshold


class StrategyBasedDetector:
    """Execute detection based on DetectionStrategy from Excel configuration
    
    This class provides detection methods that can be configured via:
    - kpi_classification_v2.xlsx: Detection_Strategy column
    - kpi_custom_rules.xlsx: Override_Strategy column
    
    Supported strategies (ivy0203 updated):
    - STANDARD_SPC: Default σ-based SPC rules
    - ABS_THRESHOLD: Absolute threshold for saturated data
    - IQR: IQR-based outlier detection for high-variance/non-normal data
    - POISSON: Poisson distribution for count/discrete data
    - RELATIVE: Relative percentage change for ratio data
    - Z_SCORE: Z-score based detection for moderate variance
    - KS_TEST: Kolmogorov-Smirnov test for distribution comparison
    - SKIP: Skip detection for constant data
    """
    
    def __init__(self, config: Optional[StrategyConfig] = None):
        self.config = config or StrategyConfig()
    
    def detect(self, strategy: DetectionStrategy,
               current_values: List[float],
               history_values: List[float],
               hist_mean: float,
               hist_std: float) -> Tuple[str, str, Dict[str, Any]]:
        """Run detection using the specified strategy
        
        Args:
            strategy: DetectionStrategy from Excel configuration
            current_values: Current data points
            history_values: Historical data points
            hist_mean: Historical mean
            hist_std: Historical std
            
        Returns:
            (severity, reason, details) tuple
            severity: 'Normal', 'Suspicious', 'Regression' or '' for fallback
        """
        # Normalize legacy strategies to core strategies
        strategy = normalize_strategy(strategy)
        
        if strategy == DetectionStrategy.SKIP:
            return 'Normal', 'Strategy: Detection skipped (configured)', {'method': 'skip'}
        
        if strategy == DetectionStrategy.IQR:
            return self._detect_by_iqr(current_values, history_values)
        
        if strategy == DetectionStrategy.POISSON:
            # Overdispersion guard: Poisson assumes Var = mean (dispersion index = 1.0).
            # When real hist_std gives dispersion >> 1, applying Poisson is wrong because
            # it uses sigma=sqrt(mean) instead of the actual (much larger) sigma, producing
            # a falsely-small p-value.  Fall back to Z_SCORE which uses the real sigma.
            # Threshold 10: if variance > 10x mean, overdispersion is severe.
            if hist_mean > 0 and hist_std > 0:
                dispersion_index = (hist_std ** 2) / hist_mean
                if dispersion_index > 10.0:
                    sev, reason, details = self._detect_by_zscore(current_values, hist_mean, hist_std)
                    details['method'] = 'poisson_overdispersed_fallback_zscore'
                    details['dispersion_index'] = round(dispersion_index, 1)
                    reason = f'Poisson overdispersed (D={dispersion_index:.0f}>>1), Z-score used: {reason}'
                    return sev, reason, details
            return self._detect_by_poisson(current_values, hist_mean)
        
        if strategy == DetectionStrategy.RELATIVE:
            return self._detect_by_relative(current_values, hist_mean)
        
        if strategy == DetectionStrategy.Z_SCORE:
            return self._detect_by_zscore(current_values, hist_mean, hist_std)
        
        if strategy == DetectionStrategy.KS_TEST:
            return self._detect_by_ks_test(current_values, history_values)
        
        # For STANDARD_SPC and ABS_THRESHOLD, return empty to use standard detection
        return '', '', {'method': 'fallback_to_standard'}
    
    def _detect_by_iqr(self, current_values: List[float],
                       history_values: List[float]) -> Tuple[str, str, Dict[str, Any]]:
        """IQR-based anomaly detection for high-variance data
        
        Uses interquartile range which is robust to outliers and non-normal distributions.
        """
        if not history_values or len(history_values) < 4:
            return 'Normal', 'IQR: Insufficient history', {}
        
        # Calculate IQR from history
        Q1 = np.percentile(history_values, 25)
        Q3 = np.percentile(history_values, 75)
        IQR = Q3 - Q1
        
        # Handle zero IQR (constant data)
        if IQR < 0.001:
            hist_median = np.median(history_values)
            curr_median = np.median(current_values)
            if abs(curr_median - hist_median) > 0.001:
                return 'Suspicious', f'IQR: Median shifted (hist={hist_median:.4f}, curr={curr_median:.4f})', {
                    'Q1': Q1, 'Q3': Q3, 'IQR': IQR, 'method': 'iqr_zero'
                }
            return 'Normal', 'IQR: Constant data', {'Q1': Q1, 'Q3': Q3, 'IQR': IQR}
        
        # Calculate bounds
        k = self.config.iqr_multiplier
        k_strict = self.config.iqr_strict_multiplier
        
        lower_bound = Q1 - k * IQR
        upper_bound = Q3 + k * IQR
        lower_strict = Q1 - k_strict * IQR
        upper_strict = Q3 + k_strict * IQR
        
        current_mean = np.mean(current_values)
        current_median = np.median(current_values)
        
        details = {
            'Q1': Q1, 'Q3': Q3, 'IQR': IQR,
            'lower_bound': lower_bound, 'upper_bound': upper_bound,
            'lower_strict': lower_strict, 'upper_strict': upper_strict,
            'current_mean': current_mean, 'current_median': current_median,
            'method': 'iqr'
        }
        
        # Check for strict violation (Regression)
        if current_mean < lower_strict or current_mean > upper_strict:
            return 'Regression', f'IQR: Mean outside 3×IQR bounds ({current_mean:.4f} vs [{lower_strict:.4f}, {upper_strict:.4f}])', details
        
        # Check for warning violation (Suspicious)
        if current_mean < lower_bound or current_mean > upper_bound:
            return 'Suspicious', f'IQR: Mean outside 1.5×IQR bounds ({current_mean:.4f} vs [{lower_bound:.4f}, {upper_bound:.4f}])', details
        
        return 'Normal', 'IQR: Within bounds', details
    
    def _detect_by_poisson(self, current_values: List[float],
                           hist_mean: float) -> Tuple[str, str, Dict[str, Any]]:
        """Poisson-based detection for count/discrete data"""
        from scipy.stats import poisson
        
        if hist_mean <= 0:
            return 'Normal', 'Poisson: Invalid λ (hist_mean <= 0)', {'method': 'poisson'}
        
        current_mean = np.mean(current_values)
        
        # Calculate two-tailed p-value
        if current_mean >= hist_mean:
            p_value = 1 - poisson.cdf(int(current_mean) - 1, mu=hist_mean)
        else:
            p_value = poisson.cdf(int(current_mean), mu=hist_mean)
        
        p_value = min(p_value * 2, 1.0)  # Two-tailed
        
        details = {
            'hist_mean_lambda': hist_mean,
            'current_mean': current_mean,
            'p_value': p_value,
            'method': 'poisson'
        }
        
        alpha = self.config.poisson_alpha
        alpha_strict = self.config.poisson_alpha_strict
        
        if p_value < alpha_strict:
            return 'Regression', f'Poisson: Significant deviation (p={p_value:.4f} < {alpha_strict})', details
        elif p_value < alpha:
            return 'Suspicious', f'Poisson: Marginal deviation (p={p_value:.4f} < {alpha})', details
        
        return 'Normal', f'Poisson: Normal (p={p_value:.4f})', details
    
    def _detect_by_relative(self, current_values: List[float],
                            hist_mean: float) -> Tuple[str, str, Dict[str, Any]]:
        """Relative change detection for ratio/percentage data"""
        current_mean = np.mean(current_values)
        
        # Handle near-zero history
        if abs(hist_mean) < 0.001:
            if abs(current_mean) > 0.001:
                return 'Suspicious', 'Relative: Value appeared from near-zero', {
                    'hist_mean': hist_mean, 'current_mean': current_mean,
                    'rel_change': float('inf'), 'method': 'relative'
                }
            return 'Normal', 'Relative: Both near zero', {'method': 'relative'}
        
        # Calculate relative change
        rel_change = abs(current_mean - hist_mean) / abs(hist_mean) * 100
        
        details = {
            'hist_mean': hist_mean,
            'current_mean': current_mean,
            'rel_change_pct': rel_change,
            'method': 'relative'
        }
        
        warning_threshold = self.config.relative_change_warning
        critical_threshold = self.config.relative_change_critical
        
        if rel_change > critical_threshold:
            direction = 'increased' if current_mean > hist_mean else 'decreased'
            return 'Regression', f'Relative: {direction} by {rel_change:.1f}% (>{critical_threshold}%)', details
        elif rel_change > warning_threshold:
            direction = 'increased' if current_mean > hist_mean else 'decreased'
            return 'Suspicious', f'Relative: {direction} by {rel_change:.1f}% (>{warning_threshold}%)', details
        
        return 'Normal', f'Relative: Change {rel_change:.1f}% within bounds', details
    
    @staticmethod
    def is_alternative_strategy(strategy: DetectionStrategy) -> bool:
        """Check if strategy is one of the alternative methods that have custom implementations"""
        # Normalize first
        strategy = normalize_strategy(strategy)
        return strategy in (DetectionStrategy.IQR, DetectionStrategy.POISSON, 
                           DetectionStrategy.RELATIVE, DetectionStrategy.Z_SCORE,
                           DetectionStrategy.KS_TEST, DetectionStrategy.SKIP)
    
    def _detect_by_zscore(self, current_values: List[float],
                          hist_mean: float, hist_std: float) -> Tuple[str, str, Dict[str, Any]]:
        """Z-score based detection for moderate variance data
        
        Uses z-score to detect deviation from historical distribution.
        """
        if hist_std <= 0:
            # Constant history - any change is significant
            current_mean = np.mean(current_values)
            if abs(current_mean - hist_mean) > 0.001:
                return 'Suspicious', f'Z-Score: Deviation from constant (hist={hist_mean:.4f}, curr={current_mean:.4f})', {
                    'method': 'zscore', 'hist_mean': hist_mean, 'hist_std': 0
                }
            return 'Normal', 'Z-Score: Constant data unchanged', {'method': 'zscore'}
        
        current_mean = np.mean(current_values)
        z_score = (current_mean - hist_mean) / hist_std
        
        details = {
            'hist_mean': hist_mean,
            'hist_std': hist_std,
            'current_mean': current_mean,
            'z_score': z_score,
            'method': 'zscore'
        }
        
        warning_threshold = self.config.zscore_warning
        critical_threshold = self.config.zscore_critical
        
        abs_z = abs(z_score)
        if abs_z > critical_threshold:
            direction = 'above' if z_score > 0 else 'below'
            return 'Regression', f'Z-Score: {abs_z:.2f}σ {direction} mean (>{critical_threshold}σ)', details
        elif abs_z > warning_threshold:
            direction = 'above' if z_score > 0 else 'below'
            return 'Suspicious', f'Z-Score: {abs_z:.2f}σ {direction} mean (>{warning_threshold}σ)', details
        
        return 'Normal', f'Z-Score: {abs_z:.2f}σ within bounds', details
    
    def _detect_by_ks_test(self, current_values: List[float],
                          history_values: List[float]) -> Tuple[str, str, Dict[str, Any]]:
        """Kolmogorov-Smirnov test for distribution comparison
        
        Compares current distribution against historical distribution.
        Useful for low-variance data where distribution shape matters.
        """
        from scipy.stats import ks_2samp
        
        if not history_values or len(history_values) < 5:
            return 'Normal', 'KS-Test: Insufficient history', {'method': 'ks_test'}
        
        if not current_values or len(current_values) < 3:
            return 'Normal', 'KS-Test: Insufficient current data', {'method': 'ks_test'}
        
        # Perform KS test
        try:
            statistic, p_value = ks_2samp(current_values, history_values)
        except Exception as e:
            return 'Normal', f'KS-Test: Error - {str(e)}', {'method': 'ks_test', 'error': str(e)}
        
        details = {
            'ks_statistic': statistic,
            'p_value': p_value,
            'n_current': len(current_values),
            'n_history': len(history_values),
            'method': 'ks_test'
        }
        
        alpha = self.config.ks_alpha
        alpha_strict = self.config.ks_alpha_strict
        
        if p_value < alpha_strict:
            return 'Regression', f'KS-Test: Distribution changed significantly (p={p_value:.4f} < {alpha_strict})', details
        elif p_value < alpha:
            return 'Suspicious', f'KS-Test: Distribution may have changed (p={p_value:.4f} < {alpha})', details
        
        return 'Normal', f'KS-Test: Distribution unchanged (p={p_value:.4f})', details


# Singleton instance for strategy-based detector
_strategy_detector_instance: Optional[StrategyBasedDetector] = None


def get_strategy_based_detector() -> StrategyBasedDetector:
    """Get singleton strategy-based detector instance"""
    global _strategy_detector_instance
    if _strategy_detector_instance is None:
        _strategy_detector_instance = StrategyBasedDetector()
    return _strategy_detector_instance
