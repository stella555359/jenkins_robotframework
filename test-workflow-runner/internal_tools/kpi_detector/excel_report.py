# -*- coding: utf-8 -*-
"""
Excel Report Generator Module (ivy0210 rewrite)
Generates Excel reports with 2 sheets:
  Sheet1: Severity Change Tracking - regression & suspicious codes with change flags
  Sheet2: All Raw Data - merged by code, aligned across history
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path
from typing import Dict, Any, Optional, List

from .config import COLORS
from .utils import generate_sparkline


class ExcelReportGenerator:
    """Generate Excel analysis reports"""
    
    def __init__(self, output_dir: Path):
        """Initialize report generator
        
        Args:
            output_dir: Directory for output files
        """
        self.output_dir = output_dir
        self._init_styles()
    
    def _init_styles(self):
        """Initialize Excel styles"""
        self.header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        self.critical_fill = PatternFill(start_color=COLORS['critical'], end_color=COLORS['critical'], fill_type="solid")
        self.suspicious_fill = PatternFill(start_color=COLORS['suspicious'], end_color=COLORS['suspicious'], fill_type="solid")
        self.warning_fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type="solid")
        self.caution_fill = PatternFill(start_color=COLORS['caution'], end_color=COLORS['caution'], fill_type="solid")
        self.normal_fill = PatternFill(start_color=COLORS['normal'], end_color=COLORS['normal'], fill_type="solid")
        self.improved_fill = PatternFill(start_color=COLORS['improved'], end_color=COLORS['improved'], fill_type="solid")
        self.worsened_fill = PatternFill(start_color=COLORS['worsened'], end_color=COLORS['worsened'], fill_type="solid")
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate(self, filename: str, analysis_results: Dict[str, Any], 
                 comparison: Optional[Dict] = None,
                 history_records: list = None,
                 current_record: Dict = None,
                 previous_record: Dict = None) -> Path:
        """Generate Excel report (ivy0210 rewrite)
        
        Args:
            filename: Base filename (without extension)
            analysis_results: Analysis results dictionary
            comparison: Comparison with history
            history_records: List of historical records
            current_record: Current analysis record
            previous_record: Previous analysis record (for severity change tracking)
            
        Returns:
            Path to generated report file
        """
        output_file = self.output_dir / f"{filename}_analysis_report.xlsx"
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        all_codes_df = analysis_results.get('all_codes', pd.DataFrame())
        consistency_df = analysis_results.get('consistency', pd.DataFrame())
        
        # Sheet 1: Severity Change Tracking
        self._create_severity_tracking_sheet(wb, all_codes_df, comparison, previous_record, history_records)
        
        # Sheet 2: All Raw Data merged by code
        self._create_raw_data_sheet(wb, consistency_df, history_records, current_record)
        
        wb.save(output_file)
        print(f"Report saved to: {output_file}")
        
        return output_file
    
    def _get_severity_fill(self, severity: str) -> PatternFill:
        """Get fill color for severity"""
        if severity == 'Regression':
            return self.critical_fill
        elif severity == 'Suspicious':
            return self.suspicious_fill
        elif severity == 'NoHistory':
            return PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        elif severity == 'Normal':
            return self.normal_fill
        return None
    
    def _get_change_flag(self, current_severity: str, prev_severity: str) -> str:
        """Calculate severity change flag between n-1 and current report"""
        severity_rank = {'Normal': 0, 'NoHistory': 1, 'Suspicious': 2, 'Regression': 3}
        curr_rank = severity_rank.get(current_severity, -1)
        prev_rank = severity_rank.get(prev_severity, -1)
        
        if not prev_severity:
            return 'NEW'
        elif curr_rank > prev_rank:
            return 'WORSENED'
        elif curr_rank < prev_rank:
            return 'IMPROVED'
        else:
            return 'UNCHANGED'
    
    def _safe_round(self, value, decimals=3):
        """Safely round a value"""
        if value is None or (isinstance(value, float) and (np.isnan(value) or np.isinf(value))):
            return ''
        try:
            return round(float(value), decimals)
        except (ValueError, TypeError):
            return ''
    
    def _create_severity_tracking_sheet(self, wb, all_codes_df: pd.DataFrame, 
                                         comparison: Optional[Dict],
                                         previous_record: Optional[Dict],
                                         history_records: list):
        """Create Sheet1: Severity Change Tracking
        
        Shows regression & suspicious codes with their key stats and severity change flags.
        """
        ws = wb.create_sheet(title='1_Severity_Tracking')
        
        if all_codes_df.empty:
            ws.cell(row=1, column=1, value='No analysis data available')
            return
        
        # Get previous severity map
        prev_severity_map = {}
        if previous_record:
            prev_severity_map = previous_record.get('code_severity', {})
        
        # Filter: Regression and Suspicious codes
        non_normal_df = all_codes_df[all_codes_df['Severity'].isin(['Regression', 'Suspicious'])].copy()
        
        # Identify removed regression/suspicious codes (improved to Normal)
        removed_codes = []
        for code, prev_sev in prev_severity_map.items():
            if prev_sev in ('Regression', 'Suspicious'):
                curr_row = all_codes_df[all_codes_df['Code'] == code]
                if not curr_row.empty:
                    curr_sev = curr_row['Severity'].iloc[0]
                    if curr_sev == 'Normal':
                        removed_codes.append({
                            'Code': code,
                            'KPI Name': curr_row['KPI Name'].iloc[0] if 'KPI Name' in curr_row.columns else '',
                            'Type': curr_row['Type'].iloc[0] if 'Type' in curr_row.columns else '',
                            'Severity': 'Normal',
                            'Reasons': f'Improved from {prev_sev}',
                            'Current_Mean': curr_row['Current_Mean'].iloc[0] if 'Current_Mean' in curr_row.columns else None,
                            'Hist_Mean': curr_row['Hist_Mean'].iloc[0] if 'Hist_Mean' in curr_row.columns else None,
                            'Sigma_Level': curr_row['Sigma_Level'].iloc[0] if 'Sigma_Level' in curr_row.columns else None,
                            'Anomaly_Count': curr_row['Anomaly_Count'].iloc[0] if 'Anomaly_Count' in curr_row.columns else 0,
                            'SPC_Zone': curr_row['SPC_Zone'].iloc[0] if 'SPC_Zone' in curr_row.columns else '',
                            'Detection_Strategy': curr_row['Detection_Strategy'].iloc[0] if 'Detection_Strategy' in curr_row.columns else '',
                            'Prev_Severity': prev_sev,
                            'Change_Flag': 'IMPROVED',
                        })
        
        # Headers
        headers = [
            'Code', 'KPI Name', 'Type', 'Severity', 'Reasons',
            'Current Mean', 'History Mean', 'Relative Change(%)',
            'Sigma Level', 'Anomaly Count', 'SPC Zone', 'Strategy',
            'Prev Severity', 'Change Flag'
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        row_num = 2
        
        # Section 1: Current Regression/Suspicious codes
        for _, row in non_normal_df.iterrows():
            code = row.get('Code', '')
            current_mean = row.get('Current_Mean')
            hist_mean = row.get('Hist_Mean')
            severity = row.get('Severity', '')
            
            # Calculate relative change
            relative_pct = ''
            if (current_mean is not None and hist_mean is not None 
                and pd.notna(current_mean) and pd.notna(hist_mean) 
                and abs(hist_mean) > 1e-10):
                relative_pct = self._safe_round(((current_mean - hist_mean) / abs(hist_mean)) * 100, 3)
            
            prev_sev = prev_severity_map.get(code, '')
            change_flag = self._get_change_flag(severity, prev_sev if prev_sev else None)
            
            row_data = [
                code,
                row.get('KPI Name', ''),
                row.get('Type', ''),
                severity,
                row.get('Reasons', ''),
                self._safe_round(current_mean),
                self._safe_round(hist_mean),
                relative_pct,
                self._safe_round(row.get('Sigma_Level')),
                int(row.get('Anomaly_Count', 0)) if pd.notna(row.get('Anomaly_Count', 0)) else 0,
                row.get('SPC_Zone', ''),
                row.get('Detection_Strategy', ''),
                prev_sev if prev_sev else 'N/A',
                change_flag,
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.border = self.thin_border
                
                if col_idx == 4:  # Severity
                    fill = self._get_severity_fill(str(value))
                    if fill:
                        cell.fill = fill
                        if value in ('Regression', 'Suspicious'):
                            cell.font = Font(bold=True, color="FFFFFF")
                
                if col_idx == 14:  # Change Flag
                    if value == 'WORSENED':
                        cell.fill = self.worsened_fill
                        cell.font = Font(bold=True)
                    elif value == 'IMPROVED':
                        cell.fill = self.improved_fill
                    elif value == 'NEW':
                        cell.fill = PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid")
                        cell.font = Font(bold=True)
            
            row_num += 1
        
        # Section 2: Improved codes (removed from regression/suspicious)
        if removed_codes:
            sep_cell = ws.cell(row=row_num, column=1, 
                              value='--- Improved to Normal (was Regression/Suspicious) ---')
            sep_cell.font = Font(bold=True, italic=True, color="006400")
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
            row_num += 1
            
            for item in removed_codes:
                row_data = [
                    item['Code'], item['KPI Name'], item['Type'], item['Severity'],
                    item['Reasons'],
                    self._safe_round(item['Current_Mean']),
                    self._safe_round(item['Hist_Mean']),
                    '', 
                    self._safe_round(item['Sigma_Level']),
                    int(item['Anomaly_Count']) if item['Anomaly_Count'] else 0,
                    item['SPC_Zone'], item['Detection_Strategy'],
                    item['Prev_Severity'], item['Change_Flag'],
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    cell.border = self.thin_border
                    if col_idx == 14:
                        cell.fill = self.improved_fill
                    if col_idx == 4:
                        cell.fill = self.normal_fill
                
                row_num += 1
        
        # Column widths
        col_widths = {
            'A': 18, 'B': 40, 'C': 10, 'D': 12, 'E': 60,
            'F': 14, 'G': 14, 'H': 16, 'I': 12, 'J': 14,
            'K': 12, 'L': 16, 'M': 14, 'N': 14
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        
        ws.freeze_panes = 'A2'
    
    def _create_raw_data_sheet(self, wb, consistency_df: pd.DataFrame,
                                history_records: list, current_record: Dict):
        """Create Sheet2: All Raw Data merged by code
        
        Rows = codes, Columns grouped by record: {record}_mean | {record}_data
        """
        ws = wb.create_sheet(title='2_All_Raw_Data')
        
        if consistency_df.empty:
            ws.cell(row=1, column=1, value='No data available')
            return
        
        # All records chronological
        all_records = list(history_records) if history_records else []
        if current_record:
            all_records.append(current_record)
        
        if not all_records:
            ws.cell(row=1, column=1, value='No records available')
            return
        
        # All unique codes
        all_codes = set()
        for rec in all_records:
            all_codes.update(rec.get('kpi_data', {}).keys())
        if 'Code' in consistency_df.columns:
            all_codes.update(consistency_df['Code'].dropna().astype(str).tolist())
        all_codes = sorted(all_codes)
        
        # Headers: Code | KPI Name | [record1_mean | record1_data] | [record2_mean | record2_data] ...
        headers = ['Code', 'KPI Name']
        for i, rec in enumerate(all_records):
            fname = rec.get('filename', f'Record_{i+1}')
            short_name = fname[:30] if len(fname) > 30 else fname
            is_current = (current_record and rec.get('filename') == current_record.get('filename'))
            label = f"{'[CUR] ' if is_current else ''}{short_name}"
            headers.append(f'{label}_mean')
            headers.append(f'{label}_data')
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border
        
        # Name lookup
        name_lookup = {}
        if not consistency_df.empty and 'Code' in consistency_df.columns:
            for _, row in consistency_df.iterrows():
                c = str(row.get('Code', '')).strip()
                if c:
                    name_lookup[c] = row.get('KPI Name', '')
        
        for row_idx, code in enumerate(all_codes, start=2):
            ws.cell(row=row_idx, column=1, value=code).border = self.thin_border
            
            kpi_name = name_lookup.get(code, '')
            if not kpi_name:
                for rec in all_records:
                    cd = rec.get('kpi_data', {}).get(code, {})
                    kpi_name = cd.get('kpi_name', '')
                    if kpi_name:
                        break
            ws.cell(row=row_idx, column=2, value=kpi_name).border = self.thin_border
            
            col_offset = 3
            for rec in all_records:
                code_data = rec.get('kpi_data', {}).get(code, {})
                mean_val = code_data.get('mean')
                raw_data = code_data.get('raw_data', '')
                
                mean_cell = ws.cell(row=row_idx, column=col_offset, value=self._safe_round(mean_val))
                mean_cell.border = self.thin_border
                
                data_cell = ws.cell(row=row_idx, column=col_offset + 1, value=raw_data)
                data_cell.border = self.thin_border
                data_cell.font = Font(size=9)
                
                is_current = (current_record and rec.get('filename') == current_record.get('filename'))
                if is_current:
                    cur_bg = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    mean_cell.fill = cur_bg
                    data_cell.fill = cur_bg
                
                col_offset += 2
        
        # Column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 40
        for i in range(len(all_records)):
            mean_col = openpyxl.utils.get_column_letter(3 + i * 2)
            data_col = openpyxl.utils.get_column_letter(4 + i * 2)
            ws.column_dimensions[mean_col].width = 14
            ws.column_dimensions[data_col].width = 50
        
        ws.freeze_panes = 'C2'
