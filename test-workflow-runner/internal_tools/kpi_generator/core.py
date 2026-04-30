from __future__ import annotations

import argparse
import ast
import json
import logging
import os
import re
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Optional, Sequence

import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ._constants import (
    COMPASS_LOGIN_RETRY_DELAY_SECONDS,
    COMPASS_LOGIN_RETRIES,
    COMPASS_LOGIN_TIMEOUT_SECONDS,
    COMPASS_REPORT_SHEET_NAME,
    DEFAULT_MAX_INTERVAL_WORKERS,
    DEFAULT_PASSWORD,
    DEFAULT_TEST_LINE_PREFIX,
    DEFAULT_USERNAME,
    DETECTOR_DATA_COLUMN_PATTERN,
    ENVIRONMENT_CODE_PATTERN,
    ENVIRONMENT_TEST_LINE_MAP,
    EXCEL_SHEET_NAME_INVALID_PATTERN,
    FINAL_REPORT_RETRIES,
    FINAL_REPORT_RETRY_DELAY_SECONDS,
    GENERATE_REPORT_RETRY_DELAY_SECONDS,
    MAX_INTERVAL_WORKERS_CAP,
    PROGRESS_PREFIX,
    SCOUT_REPORT_SHEET_NAME,
    TIMESTAMP_FORMAT,
    TEMPLATE_NAME_DELIMITER_PATTERN,
    TEMPLATE_NAME_LINEBREAK_PATTERN,
    TEMPLATE_NAME_MULTI_SPACE_PATTERN,
)
from ._text import (
    configure_logging,
    emit_progress,
    extract_environment_code,
    format_environment_filename_token,
    normalize_build,
    normalize_environment,
    normalize_template_name,
    parse_dist_name_filter_tokens,
    parse_template_name_tokens,
    parse_timestamp,
    resolve_test_line,
    sanitize_filename_token,
    split_template_name_tokens,
)


@dataclass(frozen=True)
class TimeRange:
    start: datetime
    end: datetime

    def to_strings(self) -> list[str]:
        return [self.start.strftime(TIMESTAMP_FORMAT), self.end.strftime(TIMESTAMP_FORMAT)]


@dataclass
class KpiGeneratorRequest:
    template_set_name: str
    template_names: list[str]
    build: str
    environment: str
    scenario: str
    report_timestamps_list: list[TimeRange]
    timestamp_delta_minutes: Optional[int] = None
    test_line: str = ''
    max_interval_workers: Optional[int] = None

    @classmethod
    def from_payload(cls, payload: dict[str, Any]) -> 'KpiGeneratorRequest':
        template_set_name = normalize_template_name(payload.get('template_set_name') or payload.get('release') or '')
        template_names = parse_template_name_tokens(payload.get('template_names') or '')
        if not template_set_name and not template_names:
            raise ValueError('At least one of template_set_name or template_names is required.')

        build = normalize_build(payload.get('build') or '')
        if not build:
            raise ValueError('build is required.')

        environment = normalize_environment(str(payload.get('environment') or ''))
        if not environment:
            raise ValueError('environment is required.')

        scenario = sanitize_filename_token(str(payload.get('scenario') or ''), '')
        if not scenario:
            raise ValueError('scenario is required.')

        raw_timestamps = payload.get('report_timestamps_list')
        if isinstance(raw_timestamps, str):
            report_timestamps_list = parse_report_timestamps_list(raw_timestamps)
        else:
            report_timestamps_list = coerce_report_timestamps(raw_timestamps)

        raw_delta = payload.get('timestamp_delta_minutes')
        if raw_delta in (None, '', 0, '0'):
            timestamp_delta_minutes = None
        else:
            try:
                timestamp_delta_minutes = int(raw_delta)
            except (TypeError, ValueError) as exc:
                raise ValueError('timestamp_delta_minutes must be an integer.') from exc
            if timestamp_delta_minutes <= 0:
                raise ValueError('timestamp_delta_minutes must be greater than 0 when provided.')

        resolved_test_line = resolve_test_line(environment)
        test_line = str(payload.get('test_line') or resolved_test_line).strip() or resolved_test_line

        raw_workers = payload.get('max_interval_workers')
        max_interval_workers: Optional[int] = None
        if raw_workers not in (None, '',):
            try:
                max_interval_workers = int(raw_workers)
            except (TypeError, ValueError) as exc:
                raise ValueError('max_interval_workers must be an integer.') from exc
            if max_interval_workers < 1:
                raise ValueError('max_interval_workers must be >= 1.')
            max_interval_workers = min(max_interval_workers, MAX_INTERVAL_WORKERS_CAP)

        return cls(
            template_set_name=template_set_name,
            template_names=template_names,
            build=build,
            environment=environment,
            scenario=scenario,
            report_timestamps_list=report_timestamps_list,
            timestamp_delta_minutes=timestamp_delta_minutes,
            test_line=test_line,
            max_interval_workers=max_interval_workers,
        )


@dataclass
class KpiGeneratorResult:
    template_set_name: str
    manual_template_names: list[str]
    original_intervals: list[list[str]]
    expanded_intervals: list[list[str]]
    template_names: list[str]
    combined_report_ids: list[str]
    interval_report_ids: list[str]
    final_operation: str
    report_file_path: str
    final_filename: str
    test_line: str
    build: str
    environment: str
    scenario: str
    timestamp_delta_minutes: Optional[int]
    generated_at: str
    interval_details: list[dict[str, Any]] = field(default_factory=list)
    failed_templates: list[dict[str, Any]] = field(default_factory=list)
    failed_intervals: list[dict[str, Any]] = field(default_factory=list)
    final_failure: dict[str, Any] = field(default_factory=dict)
    summary: dict[str, Any] = field(default_factory=dict)

    def to_jsonable(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class ScoutToCompassConversionResult:
    scout_report_path: str
    output_path: str
    final_filename: str
    dist_name_filter: str
    dist_name_filters: list[str]
    matched_dist_names: list[str]
    timestamp_headers: list[str]
    group_row_count: int
    build: str
    environment: str
    scenario: str
    generated_at: str
    sheet_results: list[dict[str, Any]] = field(default_factory=list)

    def to_jsonable(self) -> dict[str, Any]:
        return asdict(self)


class KpiGeneratorExecutionError(RuntimeError):
    def __init__(self, message: str, result: KpiGeneratorResult):
        super().__init__(message)
        self.result = result


class FinalReportRequestError(RuntimeError):
    def __init__(self, message: str, diagnostics: dict[str, Any]):
        super().__init__(message)
        self.diagnostics = diagnostics


def coerce_report_timestamps(raw_value: Any) -> list[TimeRange]:
    if not isinstance(raw_value, Sequence) or isinstance(raw_value, (str, bytes)):
        raise ValueError('report_timestamps_list must be a list of [start, end] pairs.')

    intervals: list[TimeRange] = []
    for item in raw_value:
        if not isinstance(item, Sequence) or isinstance(item, (str, bytes)) or len(item) != 2:
            raise ValueError('Each report_timestamps_list item must contain exactly [start_time, end_time].')
        start = parse_timestamp(str(item[0]))
        end = parse_timestamp(str(item[1]))
        if end <= start:
            raise ValueError(f'Invalid time range: {item!r}. end_time must be after start_time.')
        intervals.append(TimeRange(start=start, end=end))

    if not intervals:
        raise ValueError('report_timestamps_list cannot be empty.')
    return intervals


def parse_report_timestamps_list(value: str) -> list[TimeRange]:
    try:
        parsed = ast.literal_eval(value.strip())
    except (SyntaxError, ValueError) as exc:
        raise ValueError('report_timestamps_list must be a Python-style list, for example [["2026-01-30 03:10:44", "2026-01-30 07:40:44"]].') from exc
    return coerce_report_timestamps(parsed)


def split_intervals(intervals: Sequence[TimeRange], delta_minutes: Optional[int]) -> list[TimeRange]:
    if not delta_minutes:
        return list(intervals)

    result: list[TimeRange] = []
    delta = timedelta(minutes=delta_minutes)
    for interval in intervals:
        current = interval.start
        while current < interval.end:
            next_time = min(current + delta, interval.end)
            if (next_time - current) >= delta:
                result.append(TimeRange(start=current, end=next_time))
            current = next_time
    if not result:
        raise ValueError('timestamp_delta_minutes produced no valid split ranges.')
    return result


class CompassClient:
    LOGIN_URL = 'http://compass.dyn.nesc.nokia.net:8080/common/login_handle/'
    TEMPLATE_SET_URL = 'http://compass.dyn.nesc.nokia.net:8080/kpi/kpi_template_set/'
    GENERATE_URL = 'http://compass.dyn.nesc.nokia.net:8080/kpi/generate_kpi_report/'
    COMBINE_URL = 'http://compass.dyn.nesc.nokia.net:8080/kpi/combine_kpi_report/'
    REPORT_DETAIL_URL = 'http://compass.dyn.nesc.nokia.net:8080/kpi/get_kpi_report_detail/'
    COMPARE_URL = 'http://compass.dyn.nesc.nokia.net:8080/kpi/compare_kpi_report/'
    HORIZONTAL_MERGE_URL = 'http://compass.dyn.nesc.nokia.net:8080/kpi/horizontal_merge_reports/'

    def __init__(self, username: Optional[str] = None, password: Optional[str] = None, timeout: int = 180):
        self.username = username or os.environ.get('COMPASS_USERNAME') or DEFAULT_USERNAME
        self.password = password or os.environ.get('COMPASS_PASSWORD') or DEFAULT_PASSWORD
        if not self.username or not self.password:
            raise ValueError('Compass credentials are required via params or COMPASS_USERNAME/COMPASS_PASSWORD.')
        self.timeout = timeout
        self.session: Optional[requests.Session] = None

    def login(self) -> None:
        last_error: Optional[Exception] = None
        for attempt in range(1, COMPASS_LOGIN_RETRIES + 1):
            session = requests.session()
            try:
                emit_progress('login_attempt', f'Logging in to Compass (attempt {attempt}/{COMPASS_LOGIN_RETRIES}).', login_attempt=attempt, login_attempt_total=COMPASS_LOGIN_RETRIES)
                response = session.post(
                    url=self.LOGIN_URL,
                    data={'username': self.username, 'password': self.password},
                    verify=False,
                    timeout=COMPASS_LOGIN_TIMEOUT_SECONDS,
                )
                response.raise_for_status()
                set_cookie = response.headers.get('Set-Cookie')
                if not set_cookie:
                    raise ConnectionError('Compass login failed: missing Set-Cookie header.')
                session.headers['cookie'] = set_cookie.split(';', 1)[0]
                self.session = session
                emit_progress('login_succeeded', 'Compass login succeeded.', login_attempt=attempt)
                return
            except requests.HTTPError as exc:
                response = exc.response
                body_preview = response.text.strip()[:500] if response is not None else ''
                last_error = RuntimeError(f'Compass login failed with HTTP {response.status_code if response is not None else "?"}: {body_preview or exc}')
            except Exception as exc:  # noqa: BLE001
                last_error = exc

            if attempt < COMPASS_LOGIN_RETRIES:
                emit_progress('login_retry', f'Compass login failed on attempt {attempt}/{COMPASS_LOGIN_RETRIES}; retrying in {COMPASS_LOGIN_RETRY_DELAY_SECONDS}s.', login_attempt=attempt, login_attempt_total=COMPASS_LOGIN_RETRIES)
                time.sleep(COMPASS_LOGIN_RETRY_DELAY_SECONDS)

        raise RuntimeError(f'Compass login failed after {COMPASS_LOGIN_RETRIES} attempts: {last_error}')

    def _ensure_session(self) -> requests.Session:
        if self.session is None:
            self.login()
        return self.session

    @staticmethod
    def _normalize_report_ids(report_ids: Sequence[str]) -> list[str]:
        normalized = [str(report_id).strip().strip("'").strip('"') for report_id in report_ids if str(report_id).strip()]
        if not normalized:
            raise ValueError('report_ids cannot be empty.')
        return normalized

    @classmethod
    def _format_compare_report_ids(cls, report_ids: Sequence[str]) -> str:
        normalized = cls._normalize_report_ids(report_ids)
        return '&'.join(normalized)

    @classmethod
    def _format_report_id_list_literal(cls, report_ids: Sequence[str]) -> str:
        normalized = cls._normalize_report_ids(report_ids)
        return json.dumps(normalized)

    def list_template_sets(self) -> list[dict[str, Any]]:
        session = self._ensure_session()
        response = session.post(self.TEMPLATE_SET_URL, verify=False, timeout=60)
        response.raise_for_status()
        payload = response.json()
        data = payload.get('data') or []
        template_sets = [item for item in data if 'ivy' in str(item.get('name') or '').lower()]
        template_sets.sort(key=lambda item: str(item.get('name') or '').lower())
        return template_sets

    def get_template_set_names(self, template_set_name: str, logger: logging.Logger) -> list[str]:
        cleaned_template_set_name = normalize_template_name(template_set_name)
        if not cleaned_template_set_name:
            return []

        template_sets = self.list_template_sets()
        selected_template_set = next(
            (
                item for item in template_sets
                if normalize_template_name(item.get('name') or '') == cleaned_template_set_name
            ),
            None,
        )
        if selected_template_set is None:
            selected_template_set = next(
                (
                    item for item in template_sets
                    if normalize_template_name(item.get('name') or '').lower() == cleaned_template_set_name.lower()
                ),
                None,
            )
        if selected_template_set is None:
            raise ValueError(f'Compass KPI template set not found: {cleaned_template_set_name}.')

        template_string = str(selected_template_set.get('kpi_templates') or '')
        template_names = parse_template_name_tokens(template_string)
        if not template_names:
            raise ValueError(f'Compass template set {cleaned_template_set_name} did not contain any template names.')
        logger.info('Using %s KPI templates from template set %s.', len(template_names), cleaned_template_set_name)
        return template_names

    def generate_new_report(
        self,
        *,
        report_name: str,
        template_name: str,
        start_time: str,
        end_time: str,
        test_line: str,
        package_version: str,
        logger: logging.Logger,
        retries: int = 3,
    ) -> str:
        session = self._ensure_session()
        cleaned_template_name = normalize_template_name(template_name)
        if not cleaned_template_name:
            raise ValueError('template_name cannot be empty.')

        payload = {
            'report_name': report_name,
            'test_line': test_line,
            'package_version': package_version,
            'template_name': cleaned_template_name,
            'start_time': start_time,
            'end_time': end_time,
            'comments': '',
            'check_out_entry': 'true',
        }
        last_error: Optional[Exception] = None
        for attempt in range(1, retries + 1):
            try:
                emit_progress(
                    'template_request_started',
                    (
                        f'Calling Compass for template {cleaned_template_name} '
                        f'(attempt {attempt}/{retries}, timeout {self.timeout}s).'
                    ),
                    template_name=cleaned_template_name,
                    template_attempt=attempt,
                    template_attempt_total=retries,
                    request_timeout_seconds=self.timeout,
                )
                response = session.post(self.GENERATE_URL, data=payload, verify=False, timeout=self.timeout)
                response.raise_for_status()
                try:
                    result = response.json()
                except ValueError as exc:
                    body_preview = response.text.strip()[:500]
                    raise RuntimeError(f'Compass returned non-JSON response for template {cleaned_template_name!r}: {body_preview}') from exc
                if str(result.get('success')) == 'True':
                    report_id = str((result.get('data') or {}).get('id') or '').strip()
                    if report_id:
                        emit_progress(
                            'template_request_succeeded',
                            (
                                f'Compass accepted template {cleaned_template_name} '
                                f'on attempt {attempt}/{retries}: report_id={report_id}'
                            ),
                            template_name=cleaned_template_name,
                            template_attempt=attempt,
                            template_attempt_total=retries,
                            report_id=report_id,
                        )
                        return report_id
                error_text = str(result.get('error') or 'Compass did not return a report id.').strip()
                raise RuntimeError(error_text)
            except requests.Timeout as exc:
                last_error = RuntimeError(
                    f'Compass template request timed out after {self.timeout}s for template {cleaned_template_name!r}: {exc}'
                )
                logger.warning('Generate report timed out for template %r (attempt %s/%s): %s', cleaned_template_name, attempt, retries, last_error)
            except requests.HTTPError as exc:
                response = exc.response
                body_preview = response.text.strip()[:500] if response is not None else ''
                last_error = RuntimeError(f'HTTP {response.status_code if response is not None else "?"} for template {cleaned_template_name!r}: {body_preview or exc}')
                logger.warning('Generate report failed for template %r (attempt %s/%s): %s', cleaned_template_name, attempt, retries, last_error)
            except Exception as exc:  # noqa: BLE001
                last_error = exc
                logger.warning('Generate report failed for template %r (attempt %s/%s): %s', cleaned_template_name, attempt, retries, exc)

            if attempt < retries:
                emit_progress(
                    'template_request_retry',
                    (
                        f'Compass request failed for template {cleaned_template_name} '
                        f'(attempt {attempt}/{retries}). Retrying in {GENERATE_REPORT_RETRY_DELAY_SECONDS}s.'
                    ),
                    template_name=cleaned_template_name,
                    template_attempt=attempt,
                    template_attempt_total=retries,
                    retry_delay_seconds=GENERATE_REPORT_RETRY_DELAY_SECONDS,
                    error=str(last_error),
                )
                time.sleep(GENERATE_REPORT_RETRY_DELAY_SECONDS)

        emit_progress(
            'template_request_failed',
            (
                f'Compass request failed for template {cleaned_template_name} '
                f'after {retries} attempt(s): {last_error}'
            ),
            template_name=cleaned_template_name,
            template_attempt=retries,
            template_attempt_total=retries,
            error=str(last_error),
        )
        raise RuntimeError(f'Failed to generate Compass report for template {cleaned_template_name!r}: {last_error}')

    def combine_kpi_report(self, report_id_list: Sequence[str]) -> str:
        session = self._ensure_session()
        payload = {'report_id_list': json.dumps([str(report_id) for report_id in report_id_list])}
        response = session.post(self.COMBINE_URL, data=payload, verify=False, timeout=self.timeout)
        response.raise_for_status()
        result = response.json()
        if result.get('error'):
            raise RuntimeError(str(result['error']))
        data_text = str(result.get('data') or '').strip()
        combined_id = data_text.split(' ')[-1].strip()
        if not combined_id:
            raise RuntimeError(f'Compass combine_kpi_report returned an unexpected payload: {result}')
        return combined_id

    def get_analysis_kpi_result(
        self,
        report_id: str,
        output_path: Path,
        *,
        environment: Optional[str] = None,
        generated_at: Optional[datetime] = None,
    ) -> Path:
        session = self._ensure_session()
        response = session.post(self.REPORT_DETAIL_URL, data={'report_id': str(report_id)}, verify=False, timeout=self.timeout)
        response.raise_for_status()
        payload = response.json()
        rows = payload.get('data') or []
        if not rows:
            raise RuntimeError(f'Compass analysis detail returned no data for report_id={report_id}.')

        return write_analysis_report_workbook(
            rows=rows,
            output_path=output_path,
            environment=environment,
            generated_at=generated_at,
        )

    def get_compare_report(
        self,
        report_ids: Sequence[str],
        output_path: Path,
        *,
        environment: Optional[str] = None,
        interval_timestamps: Optional[Sequence[datetime]] = None,
        generated_at: Optional[datetime] = None,
    ) -> Path:
        return self._request_final_report(
            operation='compare',
            url=self.COMPARE_URL,
            request_data={'reports': self._format_compare_report_ids(report_ids)},
            report_ids=report_ids,
            output_path=output_path,
            fallback_sheet_name='Compare Report',
            environment=environment,
            interval_timestamps=interval_timestamps,
            generated_at=generated_at,
        )

    def horizontal_merge_reports(
        self,
        report_ids: Sequence[str],
        output_path: Path,
        *,
        environment: Optional[str] = None,
        interval_timestamps: Optional[Sequence[datetime]] = None,
        generated_at: Optional[datetime] = None,
    ) -> Path:
        return self._request_final_report(
            operation='horizontal_merge',
            url=self.HORIZONTAL_MERGE_URL,
            request_data={'report_id_list': self._format_report_id_list_literal(report_ids)},
            report_ids=report_ids,
            output_path=output_path,
            fallback_sheet_name='Horizontal Merge',
            environment=environment,
            interval_timestamps=interval_timestamps,
            generated_at=generated_at,
        )

    def _request_final_report(
        self,
        *,
        operation: str,
        url: str,
        request_data: dict[str, str],
        report_ids: Sequence[str],
        output_path: Path,
        fallback_sheet_name: str,
        environment: Optional[str] = None,
        interval_timestamps: Optional[Sequence[datetime]] = None,
        generated_at: Optional[datetime] = None,
    ) -> Path:
        normalized_report_ids = self._normalize_report_ids(report_ids)
        last_error: Optional[Exception] = None
        diagnostics: dict[str, Any] = {
            'operation': operation,
            'report_ids': normalized_report_ids,
            'url': url,
            'request_data': dict(request_data),
        }

        for attempt in range(1, FINAL_REPORT_RETRIES + 1):
            session = self._ensure_session()
            try:
                emit_progress(
                    f'final_{operation}_attempt',
                    f'Calling Compass {operation} (attempt {attempt}/{FINAL_REPORT_RETRIES}).',
                    final_operation=operation,
                    final_attempt=attempt,
                    final_attempt_total=FINAL_REPORT_RETRIES,
                    combined_report_count=len(normalized_report_ids),
                )
                response = session.post(url, data=request_data, verify=False, timeout=max(self.timeout, 300))
                response.raise_for_status()
                return self._save_response_to_path(
                    response,
                    output_path,
                    fallback_sheet_name=fallback_sheet_name,
                    operation=operation,
                    report_ids=normalized_report_ids,
                    request_data=request_data,
                    environment=environment,
                    interval_timestamps=interval_timestamps,
                    generated_at=generated_at,
                )
            except requests.HTTPError as exc:
                response = exc.response
                body_preview = response.text.strip()[:2000] if response is not None else ''
                diagnostics = {
                    **diagnostics,
                    'http_status': response.status_code if response is not None else None,
                    'response_body': body_preview,
                    'attempt': attempt,
                }
                last_error = FinalReportRequestError(
                    f'Compass {operation} failed with HTTP {response.status_code if response is not None else "?"}: {body_preview or exc}',
                    diagnostics,
                )
            except FinalReportRequestError as exc:
                diagnostics = dict(exc.diagnostics)
                last_error = exc
            except Exception as exc:  # noqa: BLE001
                diagnostics = {
                    **diagnostics,
                    'attempt': attempt,
                    'response_body': '',
                    'exception': str(exc),
                }
                last_error = FinalReportRequestError(
                    f'Compass {operation} failed: {exc}',
                    diagnostics,
                )

            if attempt < FINAL_REPORT_RETRIES:
                time.sleep(FINAL_REPORT_RETRY_DELAY_SECONDS)

        if isinstance(last_error, FinalReportRequestError):
            raise last_error
        raise FinalReportRequestError(f'Compass {operation} failed.', diagnostics)

    def _save_response_to_path(
        self,
        response: requests.Response,
        output_path: Path,
        *,
        fallback_sheet_name: str,
        operation: str,
        report_ids: Sequence[str],
        request_data: dict[str, str],
        environment: Optional[str] = None,
        interval_timestamps: Optional[Sequence[datetime]] = None,
        generated_at: Optional[datetime] = None,
    ) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        content_type = (response.headers.get('Content-Type') or '').lower()
        content_disposition = response.headers.get('Content-Disposition') or ''
        if 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type or 'attachment' in content_disposition.lower():
            output_path.write_bytes(response.content)
            if operation == 'horizontal_merge':
                normalize_horizontal_merge_report_workbook(
                    output_path,
                    environment=environment,
                    interval_timestamps=interval_timestamps,
                    generated_at=generated_at,
                )
            return output_path

        try:
            payload: Any = response.json()
        except ValueError:
            output_path.write_bytes(response.content)
            return output_path

        if isinstance(payload, dict):
            error_text = str(payload.get('error') or '').strip()
            success_value = payload.get('success')
            success_text = str(success_value).strip().lower() if success_value is not None else ''
            if error_text and success_text in {'false', '0', 'no'}:
                raise FinalReportRequestError(
                    f'Compass {operation} failed: {error_text}',
                    {
                        'operation': operation,
                        'report_ids': list(report_ids),
                        'url': response.url,
                        'request_data': dict(request_data),
                        'http_status': response.status_code,
                        'response_body': json.dumps(payload, ensure_ascii=False, indent=2),
                    },
                )

            if operation == 'compare' and isinstance(payload.get('packages'), list) and isinstance(payload.get('data'), list):
                return write_compare_report_workbook(
                    payload,
                    output_path,
                    environment=environment,
                    interval_timestamps=interval_timestamps,
                    generated_at=generated_at,
                )

        download_url = find_payload_value(payload, {'download_url', 'file_url', 'url', 'xlsx_url'})
        if isinstance(download_url, str) and download_url.strip():
            downloaded = self._ensure_session().get(download_url.strip(), verify=False, timeout=max(self.timeout, 300))
            downloaded.raise_for_status()
            output_path.write_bytes(downloaded.content)
            if operation == 'horizontal_merge':
                normalize_horizontal_merge_report_workbook(
                    output_path,
                    environment=environment,
                    interval_timestamps=interval_timestamps,
                    generated_at=generated_at,
                )
            return output_path

        candidate_report_id = find_payload_value(payload, {'report_id', 'id'})
        if fallback_sheet_name == 'Compare Report' and isinstance(candidate_report_id, (str, int)):
            return self.get_analysis_kpi_result(str(candidate_report_id), output_path)

        write_json_workbook(payload, output_path, sheet_name=fallback_sheet_name)
        return output_path


def find_payload_value(payload: Any, keys: set[str]) -> Optional[Any]:
    if isinstance(payload, dict):
        for key, value in payload.items():
            if key in keys and value not in (None, ''):
                return value
            nested = find_payload_value(value, keys)
            if nested not in (None, ''):
                return nested
    elif isinstance(payload, list):
        for item in payload:
            nested = find_payload_value(item, keys)
            if nested not in (None, ''):
                return nested
    return None


def write_json_workbook(payload: Any, output_path: Path, *, sheet_name: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31] or 'Result'

    data_rows = None
    if isinstance(payload, dict):
        if isinstance(payload.get('data'), list):
            data_rows = payload.get('data')
        elif isinstance(payload.get('results'), list):
            data_rows = payload.get('results')
    elif isinstance(payload, list):
        data_rows = payload

    if isinstance(data_rows, list) and data_rows and all(isinstance(item, dict) for item in data_rows):
        headers = sorted({key for item in data_rows for key in item.keys()})
        sheet.append(headers)
        for item in data_rows:
            sheet.append([stringify_cell(item.get(header)) for header in headers])
        style_kpi_sheet(sheet)
    else:
        sheet['A1'] = 'Raw JSON payload'
        sheet['A2'] = json.dumps(payload, ensure_ascii=False, indent=2)
        sheet['A2'].alignment = Alignment(vertical='top', wrap_text=True)
        sheet.column_dimensions['A'].width = 120
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _resolve_data_column_environment(environment: Optional[str], package_name: str) -> str:
    candidates = [environment, package_name]
    for candidate in candidates:
        if not candidate:
            continue
        try:
            return extract_environment_code(str(candidate))
        except ValueError:
            continue
    return 'T000'


def _parse_compare_package_timestamp(package_name: str) -> Optional[datetime]:
    text = str(package_name or '').strip()
    patterns = [
        r'(\d{8}_\d{6})',
        r'(\d{8}\.\d{6})',
        r'(\d{14})',
        r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})',
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        raw_value = match.group(1)
        normalized_value = raw_value.replace('.', '_')
        try:
            if re.fullmatch(r'\d{8}_\d{6}', normalized_value):
                return datetime.strptime(normalized_value, '%Y%m%d_%H%M%S')
            if re.fullmatch(r'\d{14}', raw_value):
                return datetime.strptime(raw_value, '%Y%m%d%H%M%S')
            return datetime.strptime(raw_value, TIMESTAMP_FORMAT)
        except ValueError:
            continue
    return None


def build_compare_value_column_names(
    package_list: Sequence[str],
    *,
    environment: Optional[str] = None,
    interval_timestamps: Optional[Sequence[datetime]] = None,
    generated_at: Optional[datetime] = None,
) -> list[str]:
    cleaned_package_names = [str(item or '').strip() for item in package_list]
    if len(set(cleaned_package_names)) == len(cleaned_package_names) and all(
        is_detector_compatible_data_column_name(package_name) for package_name in cleaned_package_names
    ):
        return cleaned_package_names

    normalized_names: list[str] = []
    used_names: set[str] = set()
    timestamp_fallbacks = list(interval_timestamps or [])
    base_generated_at = generated_at or datetime.now()

    for index, package_name in enumerate(cleaned_package_names):
        resolved_environment = _resolve_data_column_environment(environment, package_name)
        parsed_timestamp = _parse_compare_package_timestamp(package_name)
        if parsed_timestamp is None and index < len(timestamp_fallbacks):
            parsed_timestamp = timestamp_fallbacks[index]
        if parsed_timestamp is None:
            parsed_timestamp = base_generated_at + timedelta(seconds=index)

        column_name = f'{resolved_environment}.{parsed_timestamp.strftime("%Y%m%d_%H%M%S")}'
        while column_name in used_names:
            parsed_timestamp += timedelta(seconds=1)
            column_name = f'{resolved_environment}.{parsed_timestamp.strftime("%Y%m%d_%H%M%S")}'
        used_names.add(column_name)
        normalized_names.append(column_name)

    return normalized_names


def write_compare_report_workbook(
    payload: dict[str, Any],
    output_path: Path,
    *,
    environment: Optional[str] = None,
    interval_timestamps: Optional[Sequence[datetime]] = None,
    generated_at: Optional[datetime] = None,
) -> Path:
    package_list = [str(item or '').strip() for item in (payload.get('packages') or []) if str(item or '').strip()]
    all_data = payload.get('data') or []
    if not package_list or not isinstance(all_data, list):
        raise ValueError('Compare report payload must contain packages and data.')

    normalized_package_columns = build_compare_value_column_names(
        package_list,
        environment=environment,
        interval_timestamps=interval_timestamps,
        generated_at=generated_at,
    )

    dataframe_context: dict[str, list[Any]] = {
        'Group Name': [],
        'Code': [],
        'KPI Name': [],
    }
    for package_name in normalized_package_columns:
        dataframe_context[package_name] = []

    for row in all_data:
        if not isinstance(row, dict):
            continue
        dataframe_context['Group Name'].append(row.get('group_name'))
        dataframe_context['Code'].append(row.get('code'))
        dataframe_context['KPI Name'].append(row.get('kpi_name'))
        for index, package_name in enumerate(normalized_package_columns, start=1):
            dataframe_context[package_name].append(row.get(f'value{index}'))

    dataframe = pd.DataFrame(data=dataframe_context)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        dataframe.to_excel(writer, sheet_name=COMPASS_REPORT_SHEET_NAME, index=False)

    style_compare_report_sheet(output_path, COMPASS_REPORT_SHEET_NAME)
    return output_path


def build_analysis_value_column_name(environment: Optional[str], generated_at: Optional[datetime] = None) -> str:
    timestamp = pd.Timestamp(generated_at or datetime.now())
    if environment:
        return build_detector_data_column_name(environment, timestamp)
    return f'T000.{timestamp.strftime("%Y%m%d_%H%M%S")}'


def write_analysis_report_workbook(
    *,
    rows: Sequence[dict[str, Any]],
    output_path: Path,
    environment: Optional[str] = None,
    generated_at: Optional[datetime] = None,
) -> Path:
    value_column_name = build_analysis_value_column_name(environment, generated_at)
    dataframe_context: dict[str, list[Any]] = {
        'Group Name': [],
        'Code': [],
        'KPI Name': [],
        value_column_name: [],
    }

    for row in rows:
        if not isinstance(row, dict):
            continue
        dataframe_context['Group Name'].append(row.get('group_name') or row.get('group') or '')
        dataframe_context['Code'].append(row.get('code'))
        dataframe_context['KPI Name'].append(row.get('kpi_name'))
        dataframe_context[value_column_name].append(row.get('value'))

    dataframe = pd.DataFrame(data=dataframe_context)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        dataframe.to_excel(writer, sheet_name=COMPASS_REPORT_SHEET_NAME, index=False)

    style_compare_report_sheet(output_path, COMPASS_REPORT_SHEET_NAME)
    return output_path


def normalize_horizontal_merge_report_workbook(
    workbook_path: Path,
    *,
    environment: Optional[str] = None,
    interval_timestamps: Optional[Sequence[datetime]] = None,
    generated_at: Optional[datetime] = None,
) -> Path:
    workbook = load_workbook(workbook_path)
    if not workbook.worksheets:
        raise ValueError(f'Horizontal merge workbook has no sheets: {workbook_path}.')

    sheet = workbook[COMPASS_REPORT_SHEET_NAME] if COMPASS_REPORT_SHEET_NAME in workbook.sheetnames else workbook.worksheets[0]
    sheet.title = COMPASS_REPORT_SHEET_NAME

    if sheet.max_column >= 1:
        sheet.cell(row=1, column=1).value = 'Group Name'
    if sheet.max_column >= 2:
        sheet.cell(row=1, column=2).value = 'Code'
    if sheet.max_column >= 3:
        sheet.cell(row=1, column=3).value = 'KPI Name'

    if sheet.max_column >= 4:
        raw_column_names = [
            str(sheet.cell(row=1, column=column_index).value or f'value{column_index - 3}').strip()
            for column_index in range(4, sheet.max_column + 1)
        ]
        normalized_column_names = build_compare_value_column_names(
            raw_column_names,
            environment=environment,
            interval_timestamps=interval_timestamps,
            generated_at=generated_at,
        )
        for column_index, column_name in enumerate(normalized_column_names, start=4):
            sheet.cell(row=1, column=column_index).value = column_name

    workbook.save(workbook_path)
    style_compare_report_sheet(workbook_path, COMPASS_REPORT_SHEET_NAME)
    return workbook_path


def stringify_cell(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def style_kpi_sheet(sheet) -> None:
    header_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    body_font = Font(name='Calibri', size=10)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    preferred_widths = {'A': 32, 'B': 60, 'C': 25}
    for cell in sheet[1]:
        width = preferred_widths.get(cell.column_letter, 18)
        sheet.column_dimensions[cell.column_letter].width = width
    sheet.row_dimensions[1].height = 36


def style_compare_report_sheet(workbook_path: Path, sheet_name: str) -> None:
    workbook = load_workbook(workbook_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f'Sheet {sheet_name!r} not found in workbook {workbook_path}.')

    sheet = workbook[sheet_name]
    header_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
    header_font = Font(name='Calibri', size=10, color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    body_font = Font(name='Calibri', size=10)
    body_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = body_alignment

    sheet.column_dimensions['A'].width = 34
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 53
    for column_index in range(4, sheet.max_column + 1):
        sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = 18

    sheet.row_dimensions[1].height = 83
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 31

    workbook.save(workbook_path)


def apply_sheet_style(workbook_path: Path, sheet_name: str) -> None:
    workbook = load_workbook(workbook_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f'Sheet {sheet_name!r} not found in workbook {workbook_path}.')
    style_kpi_sheet(workbook[sheet_name])
    workbook.save(workbook_path)


def parse_entry_name(value: Any) -> tuple[str, str]:
    entry_name = str(value or '').strip()
    match = re.match(r'^\s*\[(.*?)\]\s*(.*)$', entry_name)
    if not match:
        return '', entry_name
    return match.group(1).strip(), match.group(2).strip()


def build_detector_data_column_name(environment: str, timestamp: pd.Timestamp) -> str:
    environment_code = extract_environment_code(environment)
    return f'{environment_code}.{timestamp.strftime("%Y%m%d_%H%M%S")}'


def build_interval_package_version(request: KpiGeneratorRequest, interval: TimeRange) -> str:
    detector_prefix = build_detector_data_column_name(request.environment, pd.Timestamp(interval.start))
    build_suffix = normalize_build(request.build)
    if build_suffix:
        return f'{detector_prefix}.{build_suffix}'
    return detector_prefix


def is_detector_compatible_data_column_name(value: Any) -> bool:
    return bool(DETECTOR_DATA_COLUMN_PATTERN.search(str(value or '').strip()))


def resolve_scout_sheet_name(value: Any, logger: Optional[logging.Logger] = None) -> str:
    candidate = str(value or '').strip()
    if candidate and candidate != SCOUT_REPORT_SHEET_NAME:
        message = f'Ignoring custom scout sheet name {candidate!r}; using fixed sheet {SCOUT_REPORT_SHEET_NAME!r}.'
        if logger is not None:
            logger.warning(message)
        emit_progress('scout_sheet_name_ignored', message, requested_sheet_name=candidate, effective_sheet_name=SCOUT_REPORT_SHEET_NAME)
    return SCOUT_REPORT_SHEET_NAME


def sanitize_excel_sheet_name(value: str, fallback: str = 'Sheet') -> str:
    candidate = EXCEL_SHEET_NAME_INVALID_PATTERN.sub('_', str(value or '').strip())
    candidate = re.sub(r'\s+', ' ', candidate).strip(" '")
    if not candidate:
        candidate = fallback
    return candidate[:31] or fallback


def build_unique_sheet_name(value: str, used_names: set[str]) -> str:
    base_name = sanitize_excel_sheet_name(value, 'Sheet')
    if base_name not in used_names:
        used_names.add(base_name)
        return base_name

    counter = 2
    while True:
        suffix = f'_{counter}'
        candidate = f'{base_name[: max(31 - len(suffix), 1)]}{suffix}'
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        counter += 1


def build_scout_compass_dataframe(
    *,
    dataframe: pd.DataFrame,
    dist_name_filter: str,
    environment: str,
    resolved_sheet_name: str,
) -> tuple[pd.DataFrame, list[str], list[str]]:
    normalized_filter = str(dist_name_filter or '').strip()
    required_columns = ['DistName', 'EntryName', 'GroupName', 'TimeStamp', 'Value']
    missing_columns = [column for column in required_columns if column not in dataframe.columns]
    if missing_columns:
        raise ValueError(f'Sheet {resolved_sheet_name!r} is missing required columns: {", ".join(missing_columns)}')

    filtered_dataframe = dataframe.loc[
        dataframe['DistName'].fillna('').astype(str).str.contains(normalized_filter, case=False, regex=False),
        required_columns,
    ].copy()
    if filtered_dataframe.empty:
        raise ValueError(f'No DistName matched filter {normalized_filter!r} in sheet {resolved_sheet_name!r}.')

    matched_dist_names = sorted(
        {
            str(value).strip()
            for value in filtered_dataframe['DistName'].dropna().tolist()
            if str(value).strip()
        }
    )

    entry_parts = filtered_dataframe['EntryName'].apply(parse_entry_name)
    filtered_dataframe['Group Name'] = filtered_dataframe['GroupName'].fillna('').astype(str).str.strip()
    filtered_dataframe['Code'] = entry_parts.str[0].fillna('').astype(str).str.strip()
    filtered_dataframe['KPI Name'] = entry_parts.str[1].fillna('').astype(str).str.strip()
    filtered_dataframe['TimeStamp'] = pd.to_datetime(filtered_dataframe['TimeStamp'], errors='coerce')
    if filtered_dataframe['TimeStamp'].isna().any():
        raise ValueError('Scout report contains unsupported TimeStamp values.')

    value_conflicts = (
        filtered_dataframe.groupby(['Group Name', 'Code', 'KPI Name', 'TimeStamp'])['Value']
        .nunique(dropna=False)
        .loc[lambda series: series > 1]
    )
    if not value_conflicts.empty:
        group_name, code, kpi_name, timestamp = value_conflicts.index[0]
        raise ValueError(
            f'Conflicting values found while merging DistName matches for {normalized_filter!r}: '
            f'GroupName={group_name!r}, Code={code!r}, KPI Name={kpi_name!r}, '
            f'TimeStamp={pd.Timestamp(timestamp).strftime(TIMESTAMP_FORMAT)}.'
        )

    result_dataframe = (
        filtered_dataframe
        .drop(columns=['DistName', 'EntryName', 'GroupName'])
        .drop_duplicates(subset=['Group Name', 'Code', 'KPI Name', 'TimeStamp'])
        .pivot(index=['Code', 'KPI Name', 'Group Name'], columns='TimeStamp', values='Value')
        .sort_index(axis=0)
        .sort_index(axis=1)
        .reset_index()
    )

    timestamp_columns = [column for column in result_dataframe.columns if isinstance(column, pd.Timestamp)]
    timestamp_headers = [column.strftime(TIMESTAMP_FORMAT) for column in timestamp_columns]
    result_dataframe.columns = [
        build_detector_data_column_name(environment, column) if isinstance(column, pd.Timestamp) else column
        for column in result_dataframe.columns
    ]
    detector_columns = ['Code', 'KPI Name', 'Group Name']
    data_columns = [column for column in result_dataframe.columns if column not in detector_columns]
    result_dataframe = result_dataframe[detector_columns + data_columns]
    return result_dataframe, matched_dist_names, timestamp_headers


def convert_scout_report_to_compass_format(
    *,
    scout_report_path: Path,
    dist_name_filter: str,
    environment: str,
    build: str,
    scenario: str,
    output_path: Optional[Path] = None,
    sheet_name: str = SCOUT_REPORT_SHEET_NAME,
    verbose: bool = False,
) -> ScoutToCompassConversionResult:
    logger = configure_logging(verbose=verbose)
    generated_at = datetime.now()
    resolved_sheet_name = resolve_scout_sheet_name(sheet_name, logger)
    normalized_filter = str(dist_name_filter or '').strip()
    normalized_filters = parse_dist_name_filter_tokens(dist_name_filter)
    normalized_build = normalize_build(build)
    normalized_environment = normalize_environment(environment)
    normalized_scenario = sanitize_filename_token(scenario, '')

    if not scout_report_path.exists():
        raise FileNotFoundError(f'Scout report not found: {scout_report_path}')
    if not normalized_filter or not normalized_filters:
        raise ValueError('dist_name_filter is required.')
    if not normalized_build:
        raise ValueError('build is required.')
    if not normalized_environment:
        raise ValueError('environment is required.')
    if not normalized_scenario:
        raise ValueError('scenario is required.')

    emit_progress(
        'scout_conversion_started',
        f'Reading scout report {scout_report_path.name}.',
        dist_name_filter=normalized_filter,
        dist_name_filters=normalized_filters,
        sheet_name=resolved_sheet_name,
    )
    dataframe = pd.read_excel(scout_report_path, sheet_name=resolved_sheet_name)
    dataframe.columns = [str(column or '').strip() for column in dataframe.columns]

    final_filename = build_final_filename(
        build=normalized_build,
        environment=normalized_environment,
        scenario=normalized_scenario,
        interval_count=0,
        delta_minutes=None,
        generated_at=generated_at,
        suffix_token='scout',
    )
    resolved_output_path = output_path or scout_report_path.with_name(final_filename)
    resolved_output_path.parent.mkdir(parents=True, exist_ok=True)

    used_sheet_names: set[str] = set()
    all_matched_dist_names: list[str] = []
    all_timestamp_headers: list[str] = []
    sheet_results: list[dict[str, Any]] = []
    total_group_row_count = 0

    with pd.ExcelWriter(resolved_output_path, engine='openpyxl') as writer:
        for index, current_filter in enumerate(normalized_filters, start=1):
            result_dataframe, matched_dist_names, timestamp_headers = build_scout_compass_dataframe(
                dataframe=dataframe,
                dist_name_filter=current_filter,
                environment=normalized_environment,
                resolved_sheet_name=resolved_sheet_name,
            )
            output_sheet_name = build_unique_sheet_name(current_filter, used_sheet_names)
            result_dataframe.to_excel(writer, sheet_name=output_sheet_name, index=False)
            total_group_row_count += len(result_dataframe.index)
            all_matched_dist_names.extend(name for name in matched_dist_names if name not in all_matched_dist_names)
            all_timestamp_headers.extend(header for header in timestamp_headers if header not in all_timestamp_headers)
            sheet_results.append({
                'index': index,
                'dist_name_filter': current_filter,
                'sheet_name': output_sheet_name,
                'matched_dist_names': matched_dist_names,
                'matched_dist_name_count': len(matched_dist_names),
                'timestamp_headers': timestamp_headers,
                'timestamp_count': len(timestamp_headers),
                'group_row_count': len(result_dataframe.index),
            })

    for sheet_result in sheet_results:
        apply_sheet_style(resolved_output_path, str(sheet_result['sheet_name']))

    logger.info(
        'Converted scout report %s into compass report format at %s using %s DistName filters.',
        scout_report_path,
        resolved_output_path,
        len(normalized_filters),
    )
    emit_progress(
        'scout_conversion_completed',
        f'Converted scout report to compass format: {resolved_output_path.name}',
        output_path=str(resolved_output_path),
        dist_name_filter_count=len(normalized_filters),
        matched_dist_name_count=len(all_matched_dist_names),
        timestamp_count=len(all_timestamp_headers),
        sheet_count=len(sheet_results),
        row_count=total_group_row_count,
    )
    return ScoutToCompassConversionResult(
        scout_report_path=str(scout_report_path),
        output_path=str(resolved_output_path),
        final_filename=resolved_output_path.name,
        dist_name_filter=normalized_filter,
        dist_name_filters=normalized_filters,
        matched_dist_names=all_matched_dist_names,
        timestamp_headers=all_timestamp_headers,
        group_row_count=total_group_row_count,
        sheet_results=sheet_results,
        build=normalized_build,
        environment=normalized_environment,
        scenario=normalized_scenario,
        generated_at=generated_at.strftime(TIMESTAMP_FORMAT),
    )


class KpiGeneratorService:
    def __init__(self, client: CompassClient, output_dir: Path, logger: logging.Logger):
        self.client = client
        self.output_dir = output_dir
        self.logger = logger

    def run(self, request: KpiGeneratorRequest) -> KpiGeneratorResult:
        generated_at = datetime.now()
        template_set_template_names: list[str] = []
        if request.template_set_name:
            emit_progress(
                'prepare_templates',
                f'Loading Compass template set {request.template_set_name}.',
                template_set_name=request.template_set_name,
            )
            template_set_template_names = self.client.get_template_set_names(request.template_set_name, self.logger)

        template_names = parse_template_name_tokens(template_set_template_names + request.template_names)
        if not template_names:
            raise ValueError('No template names resolved. Select a template set or enter one or more template_names.')

        if request.template_set_name and request.template_names:
            self.logger.info(
                'Using %s templates after merging template set %s with %s manually entered template name(s).',
                len(template_names),
                request.template_set_name,
                len(request.template_names),
            )
        elif request.template_names and not request.template_set_name:
            self.logger.info('Using %s manually entered template name(s).', len(template_names))

        expanded_intervals = split_intervals(request.report_timestamps_list, request.timestamp_delta_minutes)
        self.logger.info('Generating Compass reports for %s effective interval(s).', len(expanded_intervals))
        emit_progress(
            'interval_plan_ready',
            f'Loaded {len(template_names)} templates and prepared {len(expanded_intervals)} effective interval(s).',
            template_count=len(template_names),
            original_interval_count=len(request.report_timestamps_list),
            expanded_interval_count=len(expanded_intervals),
        )

        combined_report_ids: list[str] = []
        interval_report_ids: list[str] = []
        failed_templates: list[dict[str, Any]] = []
        failed_intervals: list[dict[str, Any]] = []
        interval_details: list[dict[str, Any]] = []

        n_intervals = len(expanded_intervals)
        workers = effective_interval_worker_count(request, n_intervals)
        use_parallel = n_intervals > 1 and workers > 1
        if use_parallel:
            username, password = self.client.username, self.client.password
            if not username or not password:
                self.logger.warning(
                    "Parallel interval execution requires Compass credentials; falling back to serial intervals."
                )
                use_parallel = False

        if use_parallel:
            emit_progress(
                "interval_parallel_plan",
                f"Running {n_intervals} interval(s) with up to {workers} parallel Compass session(s).",
                interval_workers=workers,
                interval_count=n_intervals,
            )
            tasks: list[tuple[Any, Any, logging.Logger, KpiGeneratorRequest, TimeRange, tuple[str, ...], int]] = []
            for index, interval in enumerate(expanded_intervals, start=1):
                tasks.append(
                    (username, password, self.logger, request, interval, tuple(template_names), index),
                )
            results_by_index: dict[
                int, tuple[Optional[str], list[str], list[dict[str, Any]], dict[str, Any]]
            ] = {}
            with ThreadPoolExecutor(max_workers=workers) as pool:
                futures = [pool.submit(_interval_worker_entry, task) for task in tasks]
                for fut in as_completed(futures):
                    idx, payload = fut.result()
                    results_by_index[idx] = payload
            for index in sorted(results_by_index):
                combined_id, generated_ids, interval_failures, interval_detail = results_by_index[index]
                interval_report_ids.extend(generated_ids)
                failed_templates.extend(interval_failures)
                interval_details.append(interval_detail)
                if combined_id is None:
                    failed_intervals.append(
                        {
                            "interval_index": index,
                            "interval_start": interval_detail.get("interval_start", ""),
                            "interval_end": interval_detail.get("interval_end", ""),
                            "reason": "All template report generations failed for this interval.",
                        }
                    )
                    continue
                combined_report_ids.append(combined_id)
        else:
            for index, interval in enumerate(expanded_intervals, start=1):
                combined_id, generated_ids, interval_failures, interval_detail = self._generate_interval_bundle(
                    request, interval, template_names, index
                )
                interval_report_ids.extend(generated_ids)
                failed_templates.extend(interval_failures)
                interval_details.append(interval_detail)
                if combined_id is None:
                    failed_intervals.append(
                        {
                            "interval_index": index,
                            "interval_start": interval.start.strftime(TIMESTAMP_FORMAT),
                            "interval_end": interval.end.strftime(TIMESTAMP_FORMAT),
                            "reason": "All template report generations failed for this interval.",
                        }
                    )
                    continue
                combined_report_ids.append(combined_id)

        if not combined_report_ids:
            raise RuntimeError('No combined report could be produced because all template generations failed across all intervals.')

        final_filename = build_final_filename(
            build=request.build,
            environment=request.environment,
            scenario=request.scenario,
            interval_count=len(expanded_intervals),
            delta_minutes=request.timestamp_delta_minutes,
            generated_at=generated_at,
        )
        report_path = self.output_dir / final_filename

        final_failure: dict[str, Any] = {}

        if len(combined_report_ids) == 1:
            final_operation = 'analysis'
        elif len(combined_report_ids) < 4:
            final_operation = 'compare'
        else:
            final_operation = 'horizontal_merge'

        try:
            if final_operation == 'analysis':
                emit_progress('finalize_analysis', 'Generating final analysis report.', final_operation=final_operation, combined_report_count=len(combined_report_ids))
                self.client.get_analysis_kpi_result(
                    combined_report_ids[0],
                    report_path,
                    environment=request.environment,
                    generated_at=generated_at,
                )
            elif final_operation == 'compare':
                emit_progress('finalize_compare', 'Generating final compare report.', final_operation=final_operation, combined_report_count=len(combined_report_ids))
                completed_interval_timestamps = [
                    parse_timestamp(str(detail.get('interval_end') or ''))
                    for detail in interval_details
                    if str(detail.get('status') or '') == 'completed' and str(detail.get('interval_end') or '').strip()
                ]
                self.client.get_compare_report(
                    combined_report_ids,
                    report_path,
                    environment=request.environment,
                    interval_timestamps=completed_interval_timestamps,
                    generated_at=generated_at,
                )
            else:
                emit_progress('finalize_horizontal_merge', 'Generating final horizontal merge report.', final_operation=final_operation, combined_report_count=len(combined_report_ids))
                completed_interval_timestamps = [
                    parse_timestamp(str(detail.get('interval_end') or ''))
                    for detail in interval_details
                    if str(detail.get('status') or '') == 'completed' and str(detail.get('interval_end') or '').strip()
                ]
                self.client.horizontal_merge_reports(
                    combined_report_ids,
                    report_path,
                    environment=request.environment,
                    interval_timestamps=completed_interval_timestamps,
                    generated_at=generated_at,
                )
        except Exception as exc:  # noqa: BLE001
            if isinstance(exc, FinalReportRequestError):
                final_failure = dict(exc.diagnostics)
            else:
                final_failure = {
                    'operation': final_operation,
                    'report_ids': combined_report_ids,
                    'response_body': '',
                    'exception': str(exc),
                }
            result = self._build_result(
                request=request,
                generated_at=generated_at,
                expanded_intervals=expanded_intervals,
                template_names=template_names,
                combined_report_ids=combined_report_ids,
                interval_report_ids=interval_report_ids,
                final_operation=final_operation,
                report_path=report_path,
                final_filename=final_filename,
                interval_details=interval_details,
                failed_templates=failed_templates,
                failed_intervals=failed_intervals,
                final_failure=final_failure,
            )
            raise KpiGeneratorExecutionError(str(exc), result) from exc

        result = self._build_result(
            request=request,
            generated_at=generated_at,
            expanded_intervals=expanded_intervals,
            template_names=template_names,
            combined_report_ids=combined_report_ids,
            interval_report_ids=interval_report_ids,
            final_operation=final_operation,
            report_path=report_path,
            final_filename=final_filename,
            interval_details=interval_details,
            failed_templates=failed_templates,
            failed_intervals=failed_intervals,
            final_failure=final_failure,
        )
        self.logger.info('Final KPI report generated: %s', report_path)
        emit_progress('completed', f'Final KPI report generated: {report_path.name}', final_operation=final_operation, report_file_path=str(report_path), final_filename=final_filename)
        return result

    def _build_result(
        self,
        *,
        request: KpiGeneratorRequest,
        generated_at: datetime,
        expanded_intervals: Sequence[TimeRange],
        template_names: Sequence[str],
        combined_report_ids: list[str],
        interval_report_ids: list[str],
        final_operation: str,
        report_path: Path,
        final_filename: str,
        interval_details: list[dict[str, Any]],
        failed_templates: list[dict[str, Any]],
        failed_intervals: list[dict[str, Any]],
        final_failure: dict[str, Any],
    ) -> KpiGeneratorResult:
        return KpiGeneratorResult(
            template_set_name=request.template_set_name,
            manual_template_names=list(request.template_names),
            original_intervals=[interval.to_strings() for interval in request.report_timestamps_list],
            expanded_intervals=[interval.to_strings() for interval in expanded_intervals],
            template_names=list(template_names),
            combined_report_ids=combined_report_ids,
            interval_report_ids=interval_report_ids,
            final_operation=final_operation,
            report_file_path=str(report_path),
            final_filename=final_filename,
            test_line=request.test_line,
            build=request.build,
            environment=request.environment,
            scenario=request.scenario,
            timestamp_delta_minutes=request.timestamp_delta_minutes,
            generated_at=generated_at.strftime(TIMESTAMP_FORMAT),
            interval_details=interval_details,
            failed_templates=failed_templates,
            failed_intervals=failed_intervals,
            final_failure=final_failure,
            summary={
                'template_count': len(template_names),
                'original_interval_count': len(request.report_timestamps_list),
                'expanded_interval_count': len(expanded_intervals),
                'successful_interval_count': len(combined_report_ids),
                'failed_interval_count': len(failed_intervals),
                'generated_report_count': len(interval_report_ids),
                'combined_report_count': len(combined_report_ids),
                'successful_template_count': len(interval_report_ids),
                'failed_template_count': len(failed_templates),
            },
        )

    def _generate_interval_bundle(
        self,
        request: KpiGeneratorRequest,
        interval: TimeRange,
        template_names: Sequence[str],
        interval_index: int,
    ) -> tuple[Optional[str], list[str], list[dict[str, Any]], dict[str, Any]]:
        return generate_interval_bundle(self.client, self.logger, request, interval, template_names, interval_index)


def build_interval_report_name(request: KpiGeneratorRequest, interval: TimeRange, interval_index: int) -> str:
    template_source_token = sanitize_filename_token(request.template_set_name, 'manual_templates') if request.template_set_name else 'manual_templates'
    build_token = sanitize_filename_token(normalize_build(request.build), 'build')
    environment_token = sanitize_filename_token(format_environment_filename_token(request.environment), 'env')
    scenario_token = sanitize_filename_token(request.scenario, 'scenario')
    start_token = interval.start.strftime('%Y%m%d.%H%M%S')
    end_token = interval.end.strftime('%Y%m%d.%H%M%S')
    return f'{template_source_token}_{build_token}_{environment_token}_{scenario_token}_slot{interval_index}_{start_token}_{end_token}'


def effective_interval_worker_count(request: KpiGeneratorRequest, interval_count: int) -> int:
    """How many parallel Compass sessions to use for interval bundles (each worker uses its own Session)."""
    if interval_count <= 1:
        return 1
    if request.max_interval_workers == 1:
        return 1
    if request.max_interval_workers is not None:
        return max(1, min(request.max_interval_workers, interval_count, MAX_INTERVAL_WORKERS_CAP))
    env_val = os.getenv("KPI_GENERATOR_MAX_INTERVAL_WORKERS", str(DEFAULT_MAX_INTERVAL_WORKERS))
    try:
        w = int(env_val)
    except ValueError:
        w = DEFAULT_MAX_INTERVAL_WORKERS
    w = max(1, min(w, MAX_INTERVAL_WORKERS_CAP))
    return max(1, min(w, interval_count))


def generate_interval_bundle(
    client: CompassClient,
    logger: logging.Logger,
    request: KpiGeneratorRequest,
    interval: TimeRange,
    template_names: Sequence[str],
    interval_index: int,
) -> tuple[Optional[str], list[str], list[dict[str, Any]], dict[str, Any]]:
    base_report_name = build_interval_report_name(request, interval, interval_index)
    interval_package_version = build_interval_package_version(request, interval)
    logger.info(
        "Processing interval %s: %s -> %s",
        interval_index,
        interval.start.strftime(TIMESTAMP_FORMAT),
        interval.end.strftime(TIMESTAMP_FORMAT),
    )
    emit_progress(
        "interval_started",
        f"Processing interval {interval_index}: {interval.start.strftime(TIMESTAMP_FORMAT)} -> {interval.end.strftime(TIMESTAMP_FORMAT)}",
        interval_index=interval_index,
        interval_start=interval.start.strftime(TIMESTAMP_FORMAT),
        interval_end=interval.end.strftime(TIMESTAMP_FORMAT),
        template_total=len(template_names),
    )

    report_ids: list[str] = []
    failed_templates: list[dict[str, Any]] = []
    for template_index, template_name in enumerate(template_names, start=1):
        full_report_name = f'{base_report_name}_{sanitize_filename_token(template_name, "template")}'
        emit_progress(
            "template_generating",
            f"Generating template {template_index}/{len(template_names)} for interval {interval_index}: {template_name}",
            interval_index=interval_index,
            template_index=template_index,
            template_total=len(template_names),
            template_name=template_name,
        )
        try:
            report_id = client.generate_new_report(
                report_name=full_report_name,
                template_name=template_name,
                start_time=interval.start.strftime(TIMESTAMP_FORMAT),
                end_time=interval.end.strftime(TIMESTAMP_FORMAT),
                test_line=request.test_line,
                package_version=interval_package_version,
                logger=logger,
            )
        except Exception as exc:  # noqa: BLE001
            failure = {
                "interval_index": interval_index,
                "interval_start": interval.start.strftime(TIMESTAMP_FORMAT),
                "interval_end": interval.end.strftime(TIMESTAMP_FORMAT),
                "template_index": template_index,
                "template_name": template_name,
                "error": str(exc),
            }
            failed_templates.append(failure)
            logger.warning(
                "Skipping failed template %s/%s for interval %s: %s",
                template_index,
                len(template_names),
                interval_index,
                exc,
            )
            emit_progress(
                "template_failed",
                f"Failed template {template_index}/{len(template_names)} for interval {interval_index}: {template_name}",
                interval_index=interval_index,
                template_index=template_index,
                template_total=len(template_names),
                template_name=template_name,
                failed_template_count=len(failed_templates),
                error=str(exc),
            )
            continue

        report_ids.append(report_id)
        emit_progress(
            "template_generated",
            f"Generated template {template_index}/{len(template_names)} for interval {interval_index}: report_id={report_id}",
            interval_index=interval_index,
            template_index=template_index,
            template_total=len(template_names),
            template_name=template_name,
            report_id=report_id,
            successful_template_count=len(report_ids),
        )

    if not report_ids:
        emit_progress(
            "interval_skipped",
            f"Interval {interval_index} skipped because all template generations failed.",
            interval_index=interval_index,
            failed_template_count=len(failed_templates),
        )
        return None, [], failed_templates, {
            "interval_index": interval_index,
            "interval_start": interval.start.strftime(TIMESTAMP_FORMAT),
            "interval_end": interval.end.strftime(TIMESTAMP_FORMAT),
            "status": "failed",
            "successful_template_count": 0,
            "failed_template_count": len(failed_templates),
            "combined_report_id": None,
        }

    emit_progress(
        "interval_combining",
        f"Combining {len(report_ids)} report ids for interval {interval_index}.",
        interval_index=interval_index,
        generated_report_count=len(report_ids),
    )
    combined_id = client.combine_kpi_report(report_ids)
    logger.info("Combined interval %s report id: %s", interval_index, combined_id)
    emit_progress(
        "interval_combined",
        f"Combined interval {interval_index} into report_id={combined_id}.",
        interval_index=interval_index,
        combined_report_id=combined_id,
    )
    return combined_id, report_ids, failed_templates, {
        "interval_index": interval_index,
        "interval_start": interval.start.strftime(TIMESTAMP_FORMAT),
        "interval_end": interval.end.strftime(TIMESTAMP_FORMAT),
        "status": "completed",
        "successful_template_count": len(report_ids),
        "failed_template_count": len(failed_templates),
        "combined_report_id": combined_id,
    }


def _interval_worker_entry(
    args: tuple[Any, Any, logging.Logger, KpiGeneratorRequest, TimeRange, tuple[str, ...], int],
) -> tuple[int, tuple[Optional[str], list[str], list[dict[str, Any]], dict[str, Any]]]:
    username, password, logger, request, interval, templates_t, interval_index = args
    client = CompassClient(username=username, password=password)
    bundle = generate_interval_bundle(client, logger, request, interval, list(templates_t), interval_index)
    return interval_index, bundle


def build_final_filename(
    *,
    build: str,
    environment: str,
    scenario: str,
    interval_count: int,
    delta_minutes: Optional[int],
    generated_at: datetime,
    suffix_token: Optional[str] = None,
) -> str:
    date_token = generated_at.strftime('exec%Y%m%dT%H%M%S')
    build_token = sanitize_filename_token(normalize_build(build), 'build')
    environment_token = sanitize_filename_token(format_environment_filename_token(environment), 'env')
    scenario_token = sanitize_filename_token(scenario, 'Scenario')
    split_token = 'nosplit' if not delta_minutes else f'delta{delta_minutes}'
    otherinfo_token = f'{interval_count}csv_{split_token}'
    if suffix_token:
        otherinfo_token = f'{otherinfo_token}_{sanitize_filename_token(suffix_token, "suffix")}'
    return f'{date_token}_{build_token}_{environment_token}_{scenario_token}_{otherinfo_token}.xlsx'


def load_request_from_path(request_json_path: Path) -> KpiGeneratorRequest:
    payload = json.loads(request_json_path.read_text(encoding='utf-8'))
    return KpiGeneratorRequest.from_payload(payload)


def parse_optional_output_path(value: Optional[str]) -> Optional[Path]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return Path(text)


def run_from_request(*, request_json_path: Path, output_dir: Path, result_json_path: Optional[Path] = None, verbose: bool = False) -> KpiGeneratorResult:
    logger = configure_logging(verbose=verbose)
    request = load_request_from_path(request_json_path)
    client = CompassClient()
    service = KpiGeneratorService(client=client, output_dir=output_dir, logger=logger)
    try:
        result = service.run(request)
    except KpiGeneratorExecutionError as exc:
        if result_json_path is not None:
            result_json_path.parent.mkdir(parents=True, exist_ok=True)
            result_json_path.write_text(json.dumps(exc.result.to_jsonable(), ensure_ascii=False, indent=2), encoding='utf-8')
        raise
    if result_json_path is not None:
        result_json_path.parent.mkdir(parents=True, exist_ok=True)
        result_json_path.write_text(json.dumps(result.to_jsonable(), ensure_ascii=False, indent=2), encoding='utf-8')
    return result


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description='Generate Compass KPI reports and convert scout reports into compass-style workbooks.')
    parser.add_argument('--request-json', type=Path)
    parser.add_argument('--output-dir', type=Path)
    parser.add_argument('--result-json', type=Path)
    parser.add_argument('--scout-report', type=Path, help='Path to a scout report workbook to convert from Chart Data into compass report format.')
    parser.add_argument('--dist-name', help='Substring filter for DistName values, for example NRCELL-101.')
    parser.add_argument('--build', help='Build used for the output detector-style filename.')
    parser.add_argument('--environment', help='Environment used for the output detector-style filename.')
    parser.add_argument('--scenario', help='Scenario used for the output detector-style filename.')
    parser.add_argument('--sheet-name', default=SCOUT_REPORT_SHEET_NAME, help=argparse.SUPPRESS)
    parser.add_argument('--output-path', help='Optional output workbook path for scout-report conversion mode.')
    parser.add_argument('--verbose', action='store_true')
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    if args.scout_report is not None:
        missing_arguments = [
            option_name
            for option_name, option_value in (
                ('--dist-name', args.dist_name),
                ('--build', args.build),
                ('--environment', args.environment),
                ('--scenario', args.scenario),
            )
            if not str(option_value or '').strip()
        ]
        if missing_arguments:
            parser.error(f'scout conversion mode requires: {", ".join(missing_arguments)}')

        result = convert_scout_report_to_compass_format(
            scout_report_path=args.scout_report,
            dist_name_filter=args.dist_name,
            build=args.build,
            environment=args.environment,
            scenario=args.scenario,
            output_path=parse_optional_output_path(args.output_path),
            sheet_name=args.sheet_name,
            verbose=args.verbose,
        )
        if args.result_json is not None:
            args.result_json.parent.mkdir(parents=True, exist_ok=True)
            args.result_json.write_text(json.dumps(result.to_jsonable(), ensure_ascii=False, indent=2), encoding='utf-8')
        print(json.dumps(result.to_jsonable(), ensure_ascii=False))
        return 0

    if args.request_json is None or args.output_dir is None:
        parser.error('standard generator mode requires both --request-json and --output-dir.')

    result = run_from_request(
        request_json_path=args.request_json,
        output_dir=args.output_dir,
        result_json_path=args.result_json,
        verbose=args.verbose,
    )
    print(json.dumps(result.to_jsonable(), ensure_ascii=False))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
